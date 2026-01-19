
import time
import subprocess
from retry import retry
import json
import requests
from requests.auth import HTTPBasicAuth
import logging
import var_gpstaxi
from var_gpstaxi import imagepath, checklistpath
import os
import shutil
from lxml import etree
import openpyxl
from playwright.async_api import Page
import asyncio
import module_gpstaxi




# from playwright.async_api import Page, TimeoutError
# from datetime import datetime
#
# async def click_and_wait_api(page: Page, click_selector: str, api_path: str, timeout: int = 15000):
#     """
#     Click selector để trigger API, chờ đúng API click.
#     - Chỉ lấy API khớp api_path, bỏ qua tất cả API tự call khác.
#     - Trả về: dict chứa url, status, data, time
#     """
#     try:
#         # Chờ API click, trong khi trigger click
#         async with page.expect_response(lambda r: api_path in r.url, timeout=timeout) as resp_info:
#             await page.click(click_selector)
#             print("chuẩn bị click2")
#             await page.wait_for_selector("#btn-refresh", state="visible")
#             await page.locator("#btn-refresh").click()
#
#             print("chuẩn bị click3")
#             await page.locator("#btn-refresh").click(force=True)
#
#             print("chuẩn bị click4")
#             refresh = page.locator("css=i#btn-refresh")
#             await refresh.evaluate("el => el.click()")
#
#             print("chuẩn bị click5")
#             el = page.locator("#btn-refresh")
#             box = await el.bounding_box()
#             await page.mouse.click(box["x"] + box["width"] / 2, box["y"] + box["height"] / 2)
#
#             print("chuẩn bị click6")
#             await page.wait_for_selector("#btn-refresh", state="visible")
#             await page.evaluate("document.querySelector('#btn-refresh').click()")
#
#
#             print("chuẩn bị click7")
#             await page.evaluate("document.querySelector('#btn-refresh')")
#
#             print("chuẩn bị click8")
#             await page.evaluate("""
#                 let el = document.getElementById('btn-refresh');
#                 if (el) el.dispatchEvent(new MouseEvent('click', { bubbles: true }));
#             """)
#             print("chuẩn bị click9")
#             await page.evaluate("document.getElementById('btn-refresh').click()")
#
#             print("chuẩn bị click10")
#             await page.evaluate("""
#                 const el = document.querySelector('#btn-refresh');
#                 if (el) {
#                     el.style.pointerEvents = 'auto';
#                     el.style.zIndex = '999999';
#                     el.click();
#                 }
#             """)
#
#         response = await resp_info.value
#         status = response.status
#         url = response.url
#         time_now = datetime.now().strftime("%H:%M:%S")
#         try:
#             data = await response.json()
#         except:
#             try:
#                 data = await response.text()
#             except:
#                 data = None
#
#     except TimeoutError:
#         status = -1
#         data = None
#         url = api_path
#         time_now = datetime.now().strftime("%H:%M:%S")
#
#     # In kết quả
#     print(f"[{time_now}] CLICK API RESULT")
#     print(f"URL   : {url}")
#     print(f"Status: {status}")
#     print(f"Data  : {data}")
#     print("=============================")
#
#     return {
#         "url": url,
#         "status": status,
#         "data": data,
#         "time": time_now
#     }





async def click_and_get_api(page: Page, click_selector: str, api_path: str, max_click: int = 20, interval: int = 1000):
    """
    Click tối đa max_click lần, check tất cả response.
    - Nếu API match api_path xuất hiện, trả về status code và data.
    - Nếu không tìm thấy sau max_click lần, trả về -1, None.
    """
    result = {"status": -1, "data": None, "url": None}
    api_found = False

    async def fetch_data(res):
        try:
            data = await res.json()
        except:
            try:
                data = await res.text()
            except:
                data = None
        return data

    def response_listener(res):
        nonlocal api_found, result
        if api_path in res.url and not api_found:
            api_found = True
            result["status"] = res.status
            result["url"] = res.url
            # Lấy data async
            asyncio.create_task(set_data(res))

    async def set_data(res):
        result["data"] = await fetch_data(res)

    page.on("response", response_listener)

    for attempt in range(1, max_click + 1):
        if api_found:
            break
        # await page.click(click_selector, force=True)
        await page.click(f"xpath={click_selector}")

        print(f"Click button lần thứ: {attempt}")
        await asyncio.sleep(interval / 1000)

    page.remove_listener("response", response_listener)

    if api_found:
        print(f"✅ API match tìm thấy: {result['url']} - Status: {result['status']}")
    else:
        print(f"⛔ Không tìm thấy API {api_path} sau {max_click} click")

    return result["status"], result["data"]




# def convert_xls_to_xlsx(xml_xls_path, xlsx_path):
#     if not os.path.exists(xml_xls_path):
#         raise FileNotFoundError(f"Không tìm thấy file .xls: {xml_xls_path}")
#
#     try:
#         # Đọc XML
#         tree = etree.parse(xml_xls_path)
#         root = tree.getroot()
#
#         ns = {
#             'ss': 'urn:schemas-microsoft-com:office:spreadsheet'
#         }
#
#         wb = openpyxl.Workbook()
#         ws = wb.active
#
#         # Lấy sheet đầu tiên trong XML
#         rows = root.xpath('//ss:Worksheet/ss:Table/ss:Row', namespaces=ns)
#
#         for r_idx, row in enumerate(rows, start=1):
#             cells = row.xpath('./ss:Cell/ss:Data', namespaces=ns)
#             for c_idx, cell in enumerate(cells, start=1):
#                 ws.cell(row=r_idx, column=c_idx).value = cell.text
#
#         wb.save(xlsx_path)
#         print(f"✔ Đã convert XML Spreadsheet 2003 → XLSX: {xlsx_path}")
#         return xlsx_path
#
#     except Exception as e:
#         raise Exception(f"Lỗi convert XML Spreadsheet 2003: {e}")



def convert_xls_to_xlsx(xml_xls_path, xlsx_path):
    if not os.path.exists(xml_xls_path):
        raise FileNotFoundError(f"Không tìm thấy file .xls: {xml_xls_path}")

    try:
        # Đọc header để nhận dạng loại file
        with open(xml_xls_path, "rb") as f:
            header = f.read(10)

        # =====================================
        # CASE 1: File thực chất là XLSX (ZIP - PK)
        # =====================================
        if header.startswith(b'PK'):
            shutil.copyfile(xml_xls_path, xlsx_path)
            print(f"✔ File là XLSX giả dạng XLS → đã copy sang .xlsx: {xlsx_path}")
            return xlsx_path

        # =====================================
        # CASE 2: XML Spreadsheet 2003
        # =====================================
        with open(xml_xls_path, "rb") as f:
            content = f.read()

        # Strip UTF-8 BOM nếu có
        if content.startswith(b'\xef\xbb\xbf'):
            content = content[3:]

        root = etree.XML(content)

        ns = {'ss': 'urn:schemas-microsoft-com:office:spreadsheet'}

        wb = openpyxl.Workbook()
        ws = wb.active

        rows = root.xpath('//ss:Worksheet/ss:Table/ss:Row', namespaces=ns)

        for r_idx, row in enumerate(rows, start=1):
            col = 1
            for cell in row.xpath('./ss:Cell', namespaces=ns):
                idx = cell.get('{urn:schemas-microsoft-com:office:spreadsheet}Index')
                if idx:
                    col = int(idx)

                data = cell.find('ss:Data', namespaces=ns)
                ws.cell(row=r_idx, column=col).value = data.text if data is not None else None
                col += 1

        wb.save(xlsx_path)
        print(f"✔ Đã convert XML Spreadsheet 2003 → XLSX: {xlsx_path}")
        return xlsx_path

    except Exception as e:
        raise Exception(f"Lỗi convert Excel: {e}")



async def download_excel(page, button_xpath, save_folder, new_filename):

    os.makedirs(save_folder, exist_ok=True)

    # Bắt sự kiện download
    async with page.expect_download() as download_info:
        await page.click(f"xpath={button_xpath}")

    download = await download_info.value

    # tên file cuối cùng
    final_path = os.path.join(save_folder, new_filename)

    # Xoá file cũ nếu có
    if os.path.exists(final_path):
        os.remove(final_path)

    # Lưu file
    await download.save_as(final_path)

    print(f"✔ File Excel đã tải về: {final_path}")

    return final_path



async def compare_excel_and_web_simple_display(page, code,  excel_path, table_xpath_column, table_xpath_data, number_row, number_data, skip_column=None):

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Header ở row 5
    excel_headers = [
        str(c.value).strip() if c.value else ""
        for c in sheet[number_row]
    ]

    # Row đầu tiên ở row 9
    excel_first_row = [
        str(c.value).strip() if c.value else ""
        for c in sheet[number_data]
    ]

    # --- WEB: LẤY TH + KIỂM TRA STYLE ẨN ---
    th_elements = page.locator(f"{table_xpath_column}")
    th_count = await th_elements.count()

    web_headers = []
    visible_column_indexes = []      # index của cột hiển thị trên web

    for i in range(th_count):
        th = th_elements.nth(i)

        # Lấy thuộc tính style
        style = await th.get_attribute("style")

        # # Nếu style="display:none" => bỏ qua
        # if style and "display:none" in style.replace(" ", "").lower():
        #     continue

        # Lấy text bình thường
        header_text = (await th.inner_text()).strip()
        web_headers.append(header_text)
        print(web_headers)
        # lưu index thật
        visible_column_indexes.append(i + 1)

    # --- WEB: dữ liệu row đầu tiên, chỉ lấy cột hiển thị ---
    web_first_row = []
    for col_idx in visible_column_indexes:
        cell_xpath = f"{table_xpath_data}[{col_idx}]"
        val = await page.locator(cell_xpath).inner_text()
        web_first_row.append(val.strip())

    # ========== SO SÁNH ==========
    print("\n===== 🔍 SO SÁNH HEADER =====")

    min_header_len = min(len(excel_headers), len(web_headers))

    for i in range(min_header_len):
        if excel_headers[i] != web_headers[i]:
            print(f"❌ Cột {i + 1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
            logging.info(f"❌ Cột {i + 1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            append_data(checklistpath, "Checklist", code, 6, f"❌ Cột {i + 1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
        else:
            print(f"✔ Cột {i + 1} OK")
            logging.info(f"✔ Cột {i + 1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}' OK")
            # writeData(checklistpath, "Checklist", code, 7, "Pass")
            # append_data(checklistpath, "Checklist", code, 6, f"✔ Cột {i + 1} OK")



    # Web nhiều cột hơn Excel
    if len(web_headers) > len(excel_headers):
        for i in range(len(excel_headers), len(web_headers)):
            if f"{web_headers[i]}" not in skip_column:
                print(f"❌ Excel thiếu cột {i + 1}: Web='{web_headers[i]}'")
                logging.info(f"❌ Excel thiếu cột {i + 1}: Web='{web_headers[i]}'")
                writeData(checklistpath, "Checklist", code, 7, "Fail")
                append_data(checklistpath, "Checklist", code, 6, f"❌ Excel thiếu cột {i + 1}: Web='{web_headers[i]}'")

    # Excel nhiều cột hơn Web
    if len(excel_headers) > len(web_headers):
        for i in range(len(web_headers), len(excel_headers)):
            print(f"❌ Web thiếu cột {i + 1}: Excel='{excel_headers[i]}'")
            logging.info(f"❌ Web thiếu cột {i + 1}: Excel='{excel_headers[i]}'")
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            append_data(checklistpath, "Checklist", code, 6, f"❌ Web thiếu cột {i + 1}: Excel='{excel_headers[i]}'")


    #check pass
    def get_datachecklist(code):
        workbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
        sheet = workbook["Checklist"]
        rownum = 9
        while rownum < 3000:
            rownum += 1
            if sheet[f"A{rownum}"].value == code:
                result = sheet[f"F{rownum}"].value
                return result
        return None

    result = get_datachecklist(code)
    if (result == "None") or (result == None):
        writeData(checklistpath, "Checklist", code, 7, "Pass")



async def compare_excel_and_web_simple(page, code,  excel_path, table_xpath_column, table_xpath_data, number_row, number_data):

    wb = openpyxl.load_workbook(excel_path)
    sheet = wb.active

    # Header ở row 5
    excel_headers = [
        str(c.value).strip() if c.value else ""
        for c in sheet[number_row]
    ]

    # Row đầu tiên ở row 9
    excel_first_row = [
        str(c.value).strip() if c.value else ""
        for c in sheet[number_data]
    ]

    # --- WEB: LẤY TH + KIỂM TRA STYLE ẨN ---
    th_elements = page.locator(f"{table_xpath_column}")
    th_count = await th_elements.count()

    web_headers = []
    visible_column_indexes = []      # index của cột hiển thị trên web

    for i in range(th_count):
        th = th_elements.nth(i)

        # Lấy thuộc tính style
        style = await th.get_attribute("style")

        # Nếu style="display:none" => bỏ qua
        if style and "display:none" in style.replace(" ", "").lower():
            continue

        # Lấy text bình thường
        header_text = (await th.inner_text()).strip()
        web_headers.append(header_text)

        # lưu index thật
        visible_column_indexes.append(i + 1)

    # --- WEB: dữ liệu row đầu tiên, chỉ lấy cột hiển thị ---
    web_first_row = []
    for col_idx in visible_column_indexes:
        cell_xpath = f"{table_xpath_data}[{col_idx}]"
        try:
            val = await page.locator(cell_xpath).inner_text(timeout=10000)
            web_first_row.append(val.strip())
        except:
            print(f"Khôgn tìm thấy xpath: {cell_xpath}")

    # ========== SO SÁNH ==========
    print("\n===== 🔍 SO SÁNH HEADER =====")
    for i in range(len(web_headers)):
        # KHÔNG in "Excel thiếu cột số…" nữa
        if i < len(excel_headers) and excel_headers[i] != web_headers[i]:
            logging.info(f"❌ Cột {i+1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
            print(f"❌ Cột {i+1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            append_data(checklistpath, "Checklist", code, 6, f"❌ Cột {i+1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}'")
        else:
            print(f"✔ Cột {i+1} OK")
            logging.info(f"✔ Cột {i+1}: Excel='{excel_headers[i]}' | Web='{web_headers[i]}' OK")
            # writeData(checklistpath, "Checklist", code, 7, "Pass")



    print("\n===== 🔍 SO SÁNH DÒNG ĐẦU TIÊN =====")
    for i in range(len(web_first_row)):
        if i < len(excel_first_row) and excel_first_row[i] != web_first_row[i]:
            print(f"❌ Ô (Row={number_row}, Col={i+1}): Excel='{excel_first_row[i]}' | Web='{web_first_row[i]}'")
            logging.info(f"❌ Ô (Row={number_row}, Col={i+1}): Excel='{excel_first_row[i]}' | Web='{web_first_row[i]}'")
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            append_data(checklistpath, "Checklist", code, 6, f"❌ Ô (Row={number_row}, Col={i+1}): Excel='{excel_first_row[i]}' | Web='{web_first_row[i]}'")

        else:
            print(f"✔ Ô (Row=1, Col={i+1}) OK")
            logging.info(f"✔ Ô (Row={number_row}, Col={i+1}): Excel='{excel_first_row[i]}' | Web='{web_first_row[i]}' OK")
            # writeData(checklistpath, "Checklist", code, 7, "Pass")


    print("\n===== ✔ HOÀN TẤT SO SÁNH =====")

    def get_datachecklist(code):
        workbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
        sheet = workbook["Checklist"]
        rownum = 9
        while rownum < 3000:
            rownum += 1
            if sheet[f"A{rownum}"].value == code:
                result = sheet[f"F{rownum}"].value
                return result
        return None

    result = get_datachecklist(code)
    if (result == "None") or (result == None):
        writeData(checklistpath, "Checklist", code, 7, "Pass")



async def write_result_web_excel(page, code, event, result, path_module, button_xpath, table_xpath_column, table_xpath_data, number_row, number_data):

    # 1. Tải file Excel (.xls)
    xls_file = await download_excel(
        page=page,
        button_xpath=button_xpath,
        save_folder=var_gpstaxi.excelpath,
        new_filename="baocao.xls"
    )

    # 2. Đổi sang .xlsx
    xlsx_file = os.path.join(var_gpstaxi.excelpath, "baocao.xlsx")
    convert_xls_to_xlsx(xls_file, xlsx_file)


    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    # 3. So sánh dữ liệu Excel ↔ Web
    await compare_excel_and_web_simple(
        page=page,
        code=code,
        excel_path=xlsx_file,
        table_xpath_column=table_xpath_column,
        table_xpath_data=table_xpath_data,
        number_row=number_row,
        number_data=number_data
    )



async def write_result_web_excel_display(page, code, event, result, path_module, button_xpath, table_xpath_column, table_xpath_data, number_row, number_data, skip_column=None):

    # 1. Tải file Excel (.xls)
    xls_file = await download_excel(
        page=page,
        button_xpath=button_xpath,
        save_folder=var_gpstaxi.excelpath,
        new_filename="baocao.xls"
    )

    # 2. Đổi sang .xlsx
    xlsx_file = os.path.join(var_gpstaxi.excelpath, "baocao.xlsx")
    convert_xls_to_xlsx(xls_file, xlsx_file)


    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    # 3. So sánh dữ liệu Excel ↔ Web
    await compare_excel_and_web_simple_display(
        page=page,
        code=code,
        excel_path=xlsx_file,
        table_xpath_column=table_xpath_column,
        table_xpath_data=table_xpath_data,
        number_row=number_row,
        number_data=number_data,
        skip_column=skip_column
    )



async def write_result_web_excel_dowload(page, code, event, result, path_module, button_xpath, name_image):

    # 1. Tải file Excel (.xls)
    xls_file = await download_excel(
        page=page,
        button_xpath=button_xpath,
        save_folder=var_gpstaxi.excelpath,
        new_filename="baocao.xls"
    )

    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        xlsx_file = os.path.join(var_gpstaxi.excelpath, "baocao.xlsx")
        convert_xls_to_xlsx(xls_file, xlsx_file)
        logging.info("Pass")
        writeData(checklistpath, "Checklist", code, 7, "Pass")
    except Exception as e:
        logging.info(e)
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        logging.info("Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")



def print_final_report(report, start_time):
    print("\n" + "="*80)
    print("📘  BÁO CÁO TEST TỔNG HỢP (UNITTEST STYLE)")
    print("="*80)

    total_cases = 0
    total_pass = 0
    total_fail = 0

    for chrome, data in report.items():
        print(f"\n🟦 {chrome}")
        print("-" * 50)

        for case in data["cases"]:
            total_cases += 1
            if case["status"] == "PASS":
                total_pass += 1
                status = "PASSED"
            else:
                total_fail += 1
                status = "FAILED"

            print(f"- {case['name']:<20} {status:<8} {case['time']:.2f}s")

        print(f">>> ⏱️  Tổng thời gian {chrome}: {data['total_time']:.2f}s")

    print("\n" + "="*50)
    total_time = time.time() - start_time
    print(f"📊 Tổng số case: {total_cases}")
    print(f"✅ Pass: {total_pass}")
    print(f"❌ Fail: {total_fail}")
    print(f"⏱️ Tổng thời gian chạy toàn bộ: {total_time:.2f}s")
    print("="*50 + "\n")



def timerun():
    while True:
        time.sleep(3)
        timerun = time.strftime("%H:%M:%S", time.localtime())
        print("Thời gian hiện tại:", timerun)
        print("Thời gian chạy tool:", var_gpstaxi.timerun)
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 47, 2, timerun)
        if var_gpstaxi.timerun == "":
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 47, 2, timerun)
        else:
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 47, 2, var_gpstaxi.timerun)


        if var_gpstaxi.timerun == time.strftime("%H:%M", time.localtime()):
            break
        if var_gpstaxi.timerun == "":
            break



def clear_log():
    logging.basicConfig(handlers=[logging.FileHandler(filename=var_gpstaxi.logpath,
                                                      encoding='utf-8', mode='w')],  # mode='a+', w
                        format="%(asctime)s %(name)s:%(levelname)s:%(message)s",
                        datefmt="%F %A %T",
                        level=logging.INFO)




def clearData(file, sheetName, ketqua, trangthai, tenanh):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 9
    while (i < 1000):
        i += 1
        i = str(i)
        sheet["F"+i] = ketqua
        sheet["G"+i] = trangthai
        sheet["M"+i] = tenanh
        i = int(i)
    wordbook.save(file)




def clearData_luutamthoi(file,sheetName, api, web, popup):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 6
    while (i < 37):
        i += 1
        i = str(i)
        sheet["B"+i] = api
        sheet["C"+i] = web
        sheet["D"+i] = popup
        i = int(i)
    wordbook.save(file)



def clearData_luutamthoi_from(file,sheetName, api, web, popup, row_start, row_end):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = row_start
    while (i < row_end):
        i += 1
        i = str(i)
        sheet["B"+i] = api
        sheet["C"+i] = web
        sheet["D"+i] = popup
        i = int(i)
    wordbook.save(file)



def getRowCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_row)


def getColumnCount(file, sheetName):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return (sheet.max_column)


def readData(file,sheetName,rownum,columnno):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rownum,column=columnno).value


def writeData(file,sheetName,caseid,columnno,data):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 0
    while (i < 5000):
        i += 1
        i = str(i)
        if sheet["A"+i].value == caseid:
            i = int(i)
            sheet.cell(row=i, column=columnno).value = data
            break
        i = int(i)
    wordbook.save(file)


def append_data(file, sheetName, caseid, columnno, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]

    # Tìm dòng chứa caseid
    target_row = None
    for i in range(1, sheet.max_row + 1):
        if sheet[f"A{i}"].value == caseid:
            target_row = i
            break

    if target_row is None:
        print(f"❌ Không tìm thấy caseid: {caseid}")
        return

    # Lấy nội dung cũ của ô
    old_val = sheet.cell(row=target_row, column=columnno).value

    # Nếu ô đang trống → ghi luôn
    if old_val is None:
        sheet.cell(row=target_row, column=columnno).value = data
    else:
        # Nếu ô đã có dữ liệu → nối thêm xuống dòng
        sheet.cell(row=target_row, column=columnno).value = f"{old_val}\n{data}"

    workbook.save(file)
    print(f"✔ Đã append vào ô row={target_row}, col={columnno}")


def clearData_luutamthoi2(file,sheetName, column1, column2, column3, column4, column5):
    wordbook = openpyxl.load_workbook(file)
    sheet = wordbook.get_sheet_by_name(sheetName)
    i = 89
    while (i < 100):
        i += 1
        i = str(i)
        sheet["B"+i] = column1
        sheet["C"+i] = column2
        sheet["D"+i] = column3
        sheet["E"+i] = column4
        sheet["F"+i] = column5
        i = int(i)
    wordbook.save(file)






def viber_send_text():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    module_gpstaxi.check_casenone()
    module_gpstaxi.change_casenone()
    module_gpstaxi.check_casefail()
    module_gpstaxi.check_casepass()

    mucnghiemtrong = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 65, 2))
    tong_case_trong = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 66, 2))

    case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 77, 2))
    case_pass = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 87, 2))

    thoigianbatdauchay = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi , 'Sheet1', 47, 2))

    # if case_fail 0:
    time_string1 = time.strftime("%d/%m/%Y, ", time.localtime())
    time_string1 = str(time_string1)
    time_string2 = time.strftime("%H:%M", time.localtime())
    time_string2 = str(time_string2)
    print("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                              "\n- LinkTest: " + var_gpstaxi.linktest+
                                              "\n- ModeTest: " + var_gpstaxi.modetest+
                                              "\n- Số case Pass: " + case_pass+
                                              "\n- Số case False: "+ case_fail+
                                              "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)

    AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    FROM_USER_ID = "39wUHCI+06GzSnqRCl5J8Q=="

    if "taxi.binhanhcorp.com" in var_gpstaxi.linktest:
        if int(case_fail) >= 1:
            AUTH_TOKEN = "55d2a6233831c839-912a978ffdf73bf4-1402f2d88308a203"  # id Cảnh báo Autotest GPS TAXI
            FROM_USER_ID = "5KDYjRDZMXFiXEcFIjCnnQ=="


    # AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    # FROM_USER_ID = "39wUHCI+06GzSnqRCl5J8Q=="



    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    print(AUTH_TOKEN)
    print(FROM_USER_ID)


    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return


    # 2. Gửi tin nhắn văn bản
    send_url = "https://chatapi.viber.com/pa/post"
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,
        "type": "text",
        "text": ("- DateTest : "+time_string1+""+thoigianbatdauchay+" - "+time_string2+
                                                      "\n- LinkTest: " + var_gpstaxi.linktest+
                                                      "\n- ModeTest: " + var_gpstaxi.modetest+
                                                      "\n- Số case Pass: " + case_pass+
                                                      "\n- Số case False: "+ case_fail+
                                                      "\n- Số case False nghiêm trọng: "+ mucnghiemtrong)}
    headers = {
        "Content-Type": "application/json"}
    response = requests.post(send_url, json=payload, headers=headers)

    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())



def check_user_id():
    res = requests.post(
        "https://chatapi.viber.com/pa/get_account_info",
        headers={"X-Viber-Auth-Token": "55d2a6233831c839-912a978ffdf73bf4-1402f2d88308a203"}
    )
    print(res.json())


    # AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    # FROM_USER_ID = "39wUHCI+06GzSnqRCl5J8Q=="

    # AUTH_TOKEN = "55d2a6233831c839-912a978ffdf73bf4-1402f2d88308a203"  # id Cảnh báo Autotest GPS TAXI
    # FROM_USER_ID = "5KDYjRDZMXFiXEcFIjCnnQ=="



def upload_pixeldrain_auth(filepath):
    API_KEY = "c567bb13-f4c0-4aac-b9bd-c8add1e467fc"  # Thay bằng key thật

    with open(filepath, "rb") as f:
        res = requests.post(
            "https://pixeldrain.com/api/file",
            files={"file": f},
            auth=HTTPBasicAuth('', API_KEY)
        )
        res_json = json.loads(res.text)
        file_id = res_json["id"]
        link_download = (f"https://pixeldrain.com/api/file/{file_id}")
        print(link_download)
        return link_download



def send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, file_path):
    file_url = upload_pixeldrain_auth(file_path)
    if not file_url:
        print("⚠️ Không thể upload file. Hủy gửi.")
        return

    file_name = os.path.basename(file_path)
    file_size = os.path.getsize(file_path)


    # 1. Thiết lập webhook (tạm thời, có thể dùng URL giả nếu không cần nhận sự kiện)
    webhook_url = "https://eoj9bp6x8fvrpv8.m.pipedream.net"  # Hoặc URL server thực tế nếu có

    webhook_response = requests.post(
        "https://chatapi.viber.com/pa/set_webhook",
        headers={"X-Viber-Auth-Token": AUTH_TOKEN},
        json={"url": webhook_url})

    if webhook_response.json().get("status") != 0:
        print("⚠️ Không thể thiết lập webhook. Hủy gửi tin nhắn.")
        return

    # 2. Gửi tin nhắn văn bản
    payload = {
        "auth_token": AUTH_TOKEN,
        "from": FROM_USER_ID,           # Viber user ID người nhận
        "min_api_version": 1,
        "tracking_data": "",               # Có thể để chuỗi rỗng nếu không dùng tracking
        "type": "file",
        "media": file_url,
        "size": file_size,
        "file_name": file_name}

    headers = {
        "X-Viber-Auth-Token": AUTH_TOKEN,
        "Content-Type": "application/json"}

    response = requests.post("https://chatapi.viber.com/pa/post", json=payload, headers=headers)
    print("📨 Phản hồi từ Viber:", response.status_code, response.json())
    print("\n== Send Message Response ==")
    print("Status Code:", response.status_code)
    print("Response:", response.json())



def viber_send_file():
    # ==== Ví dụ sử dụng ====
    AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    FROM_USER_ID = "39wUHCI+06GzSnqRCl5J8Q=="
    case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 77, 2))
    if "taxi.binhanhcorp.com" in var_gpstaxi.linktest:
        if int(case_fail) >= 1:
            AUTH_TOKEN = "55d2a6233831c839-912a978ffdf73bf4-1402f2d88308a203"  # id Cảnh báo Autotest GPS TAXI
            FROM_USER_ID = "5KDYjRDZMXFiXEcFIjCnnQ=="

    # AUTH_TOKEN = "54c3dd27a1b404ea-12a24f08fb9f6d31-2de1680f6bbb4010"  # id Group QA Test
    # FROM_USER_ID = "39wUHCI+06GzSnqRCl5J8Q=="
    #
    #
    # AUTH_TOKEN = "54c41d657531134a-bb80f737c680cad2-dbfb089a1539d6bd"  # id Cảnh báo Autotest BA_GPS V3
    # FROM_USER_ID = "IPxM6i4G19VsehpqFWyY/w=="

    FILE_PATH_checklisst = var_gpstaxi.checklistpath  # Thay bằng đường dẫn file thật
    FILE_PATH_log = var_gpstaxi.logpath  # Thay bằng đường dẫn file thật

    print(var_gpstaxi.checklistpath)
    print(var_gpstaxi.logpath)


    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_checklisst)
    send_gofile_link_via_viber(AUTH_TOKEN, FROM_USER_ID, FILE_PATH_log)


def send_viber():
    viber_send_text()
    viber_send_file()













def delete_image():
    path = os.path.join(var_gpstaxi.imagepath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))




def delete_excel():
    path = os.path.join(var_gpstaxi.excelpath)
    l = os.listdir(path)
    print(l)
    for i in l:
        print(os.path.join(path, i))
        os.remove(os.path.join(path, i))



def get_datachecklist(ma):
        wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
        sheet = wordbook.get_sheet_by_name("Checklist")
        rownum = 9
        while (rownum < 3000):
            rownum += 1
            rownum = str(rownum)
            if sheet["A"+rownum].value == ma:
                tensukien = sheet["B"+rownum].value
                ketqua = sheet["E"+rownum].value
                print(ma)
                print(tensukien)
                print(ketqua)
                logging.info("đang chạy case: " + ma)
            rownum = int(rownum)



@retry(tries=2, delay=2, backoff=1, jitter=5)
def switch_tab_0(page):
    try:
        # Nghỉ tạm 1s (Playwright không có implicit wait)
        time.sleep(1)
    except Exception:
        var_gpstaxi.restart_driver()

    # --- Xử lý alert/dialog nếu có ---
    try:
        def handle_dialog(dialog):
            dialog.accept()
        var_gpstaxi.page.on("dialog", handle_dialog)
        time.sleep(1)
    except Exception:
        pass

    # --- Chạy file cancel.exe (nếu có) ---
    try:
        subprocess.Popen(var_gpstaxi.uploadpath + "cancel.exe")
    except Exception:
        pass

    # --- Đóng các tab phụ, chỉ giữ lại tab đầu tiên ---
    try:
        pages = var_gpstaxi.page.context.pages
        if len(pages) > 0:
            main_page = pages[0]
            for p in pages[1:]:
                p.close()  # Đóng các tab khác
            var_gpstaxi.page = main_page
            var_gpstaxi.page.bring_to_front()
            time.sleep(1)
    except Exception:
        pass

    # --- Khởi tạo lại trang chính ---
    try:
        var_gpstaxi.restart_driver()
        page = var_gpstaxi.context.new_page()
        page.goto(var_gpstaxi.linktest)
        page.wait_for_load_state("load", timeout=10000)
        time.sleep(2)
    except Exception:
        pass
















async def write_result_text_inner_text(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        locator = page.locator(f"xpath={path_text}")
        check_text = await locator.inner_text(timeout=1500)
        logging.info(f"Text: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            logging.info("Pass")
            writeData(checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("Fail")
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")



async def write_result_text_inner_text_in(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        locator = page.locator(f"xpath={path_text}")
        check_text = await locator.inner_text(timeout=1500)
        logging.info(f"Text: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_result in check_text:
            logging.info("Pass")
            writeData(checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("Fail")
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


# logging.info("Giám sát - Danh sách cảnh báo truyền C08 Bộ Công an")
# logging.info(f"Mã - {code}")
# logging.info(f"Tên sự kiện - {event}")
# logging.info(f"Kết quả - {result}")
# try:
#     stt = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_1}").inner_text()
#     number = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_2}").inner_text()
#     liscens_plate = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_3}").inner_text()
#     conten = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_4}").inner_text()
#     day = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_5}").inner_text()
#     person = await self.page.locator(f"xpath={var_gpstaxi.list_bca_1_6}").inner_text()
#
#     logging.info(f"STT: {stt}\nSố hiệu xe: {number}\nBiển số xe: {liscens_plate}\n"
#                  f"Nội dung cảnh báo: {conten}\nNgày bỏ tích truyền: {day}\nNgười bỏ tích truyền: {person}")
#
#     module_other_gpstaxi.writeData(checklistpath, "Checklist", code, 6, f"STT: {stt}\nSố hiệu xe: {number}\nBiển số xe: {liscens_plate}\n"
#                  f"Nội dung cảnh báo: {conten}\nNgày bỏ tích truyền: {day}\nNgười bỏ tích truyền: {person}")
#
#     if (stt == "1") and (number != "") and (liscens_plate != "") and (conten != ""):
#         logging.info("Pass")
#         module_other_gpstaxi.writeData(checklistpath, "Checklist", code, 7, "Pass")
#     else:
#         logging.info("Fail")
#         await self.page.screenshot(path=f"{imagepath}{code}_C08BCA_CheckThongTin.png", full_page=True)
#         module_other_gpstaxi.writeData(checklistpath, "Checklist", code, 7, "Fail")
#         module_other_gpstaxi.writeData(checklistpath, "Checklist", code, 13, f"{code}_C08BCA_CheckThongTin.png")
# except Exception as e:
#     logging.info(f"Fail - {e}")
#     await self.page.screenshot(path=f"{imagepath}{code}_C08BCA_CheckThongTin.png", full_page=True)
#     module_other_gpstaxi.writeData(checklistpath, "Checklist", code, 7, "Fail")




async def write_result_text_inner_text_other(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        locator = page.locator(f"xpath={path_text}")
        await locator.wait_for(timeout=1500)
        check_text = await locator.text_content()

        logging.info(f"Text: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_text != check_result:
            logging.info("Pass")
            writeData(checklistpath, "Checklist", code, 7, "Pass")
        else:
            logging.info("Fail")
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_text_text_content(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        locator = page.locator(f"xpath={path_text}")
        await locator.wait_for(timeout=1500)
        check_text = await locator.text_content()
        logging.info(f"TextContent: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_text == check_result:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            logging.info("Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        logging.info("Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_text_text_content_other(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        locator = page.locator(f"xpath={path_text}")
        check_text = await locator.text_content(timeout=1500)
        logging.info(f"TextContent: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_text != check_result:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            logging.info("Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        logging.info("Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_text_content_handle(page, code, event, result, path_module, link, path_check, desire, name_image):
    async with page.context.expect_page() as new_page_info:
        await page.click(f"xpath={link}")

    # LẤY TAB MỚI (phải có await)
    new_page = await new_page_info.value

    # Đợi tab mới load
    await new_page.wait_for_load_state("load", timeout=15000)

    # Lấy locator
    locator = new_page.locator(f"xpath={path_check}").first
    await locator.wait_for(state="visible", timeout=10000)

    # Lấy text
    check_text = await locator.text_content()
    print(f"{check_text}")

    # Ghi kết quả vào Excel (sync function, không cần await)
    writeData(var_gpstaxi.checklistpath, "Checklist", code, 6, check_text)

    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    if desire in check_text:
        writeData(var_gpstaxi.checklistpath, "Checklist", code, 7, "Pass")
        logging.info("Pass")
    else:
        await new_page.screenshot(path=f"{var_gpstaxi.imagepath}{code}{name_image}.png", full_page=True)
        writeData(var_gpstaxi.checklistpath, "Checklist", code, 7, "Fail")
        logging.info("Fail")
        writeData(var_gpstaxi.checklistpath, "Checklist", code, 13, code + name_image)

    await new_page.close()

    # await module_other_gpstaxi.write_result_text_content_handle(self.page, code, event, result, "Giám sát - Danh sách cảnh báo truyền C08 Bộ Công an",
    #                                                             link, path_check, desire, name_image)


async def write_result_text_content_handle_value(page, code, event, result, path_module, link, path_check, desire, name_image):
    async with page.context.expect_page() as new_page_info:
        await page.click(f"xpath={link}")

    # LẤY TAB MỚI (phải có await)
    new_page = await new_page_info.value

    # Đợi tab mới load
    await new_page.wait_for_load_state("load", timeout=15000)

    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        locator = new_page.locator(f"xpath={path_check}")
        check_text = await locator.input_value(timeout=1500)
        logging.info(f"TextContent: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if desire in check_text:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")

    await new_page.close()

    # await module_other_gpstaxi.write_result_text_content_handle(self.page, code, event, result, "Giám sát - Danh sách cảnh báo truyền C08 Bộ Công an",
    #                                                             link, path_check, desire, name_image)



async def write_result_text_content_handle_title(page, code, event, result, path_module, link, desire, name_image):
    async with page.context.expect_page() as new_page_info:
        await page.click(f"xpath={link}")

    # Lấy TAB mới
    new_page = await new_page_info.value

    # Đợi tab mới load
    await new_page.wait_for_load_state("load", timeout=15000)

    # ❗ Lấy title tab mới
    new_title = await new_page.title()
    logging.info(f"TITLE TAB MỚI: {new_title}")

    # ❗ Lấy URL tab mới
    current_url = new_page.url
    logging.info(f"URL TAB MỚI: {current_url}")

    # ---------- BẮT ĐẦU XỬ LÝ ----------
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        writeData(checklistpath, "Checklist", code, 6, new_title)

        if desire in new_title:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await new_page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await new_page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")

    await new_page.close()





async def write_result_value_other(page, code, event, result, path_module, path_text, check_result, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        locator = page.locator(f"xpath={path_text}")
        check_text = await locator.input_value(timeout=1500)
        logging.info(f"TextContent: {check_text}")
        writeData(checklistpath, "Checklist", code, 6, check_text)

        if check_text != check_result:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")






async def write_result_status_code(page, code, event, result, path_module, path_button, name_api, name_image):
    status, data = await click_and_get_api(
        page=page,
        click_selector=f"{path_button}",
        api_path=f"{name_api}",
        max_click=20,
        interval=1000
    )

    print("API status:", status)
    print("API data:", data)

    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    logging.info(f"Status code: {status}")
    logging.info(f"data: {data}")
    writeData(checklistpath, "Checklist", code, 6, f"Status code: {status}")

    if status == 200:
        logging.info("Pass")
        writeData(checklistpath, "Checklist", code, 7, "Pass")
    else:
        logging.info("Fail")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_close(page, code, event, result, path_module, path_button, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    await page.click(f"xpath={path_button}")
    await asyncio.sleep(2)

    try:
        button = page.locator(f"xpath={path_button}")
        await button.click(timeout=500)
        logging.info("Fail")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except:
        logging.info("Pass")
        writeData(checklistpath, "Checklist", code, 7, "Pass")








async def write_result_text_url_in(page, code, event, result, path_module, desire, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        current_url = page.url
        logging.info(f"URL hiện tại: {current_url}")
        writeData(checklistpath, "Checklist", code, 6, current_url)

        if desire in current_url:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
            logging.info("Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_displayed(page, code, event, result, path_module, path_text, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")

    try:
        visible = await page.locator(f"xpath={path_text}").is_visible(timeout=1500)
        logging.info(f"XPath: {path_text}, Visible: {visible}")
        if visible:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")


async def write_result_not_displayed(page, code, event, result, path_module, path_text, name_image):
    logging.info(path_module)
    logging.info(f"Mã - {code}")
    logging.info(f"Tên sự kiện - {event}")
    logging.info(f"Kết quả - {result}")
    try:
        visible = await page.locator(f"xpath={path_text}").is_visible(timeout=1500)
        logging.info(f"XPath: {path_text}, Visible: {visible}")
        if not visible:
            writeData(checklistpath, "Checklist", code, 7, "Pass")
        else:
            await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
            writeData(checklistpath, "Checklist", code, 7, "Fail")
            writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")
    except Exception as e:
        logging.info(f"Fail - {e}")
        await page.screenshot(path=f"{imagepath}{code}{name_image}.png", full_page=True)
        writeData(checklistpath, "Checklist", code, 7, "Fail")
        writeData(checklistpath, "Checklist", code, 13, f"{code}{name_image}")












