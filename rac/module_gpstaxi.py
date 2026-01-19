import time
import openpyxl
from selenium.webdriver.common.by import By
import module_other_gpstaxi
import var_gpstaxi
import re
import caseid_gpstaxi
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from collections import defaultdict



def check_casenone():
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 60, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 61, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 62, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 63, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 64, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 65, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_gpstaxi.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 61, 2, count)


        if i == "2":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 62, 2, count)


        if i == "3":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 63, 2, count)


        if i == "4":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 64, 2, count)


        if i == "0":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 66, 2, count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 60, 2, count)

        else:
            muc1 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 61, 2))
            muc2 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 62, 2))
            muc3 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 63, 2))
            muc4 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 64, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 60, 2, sumarry_case_none)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 66, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1350):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 65, 2, count)



def check_casefail():
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_gpstaxi.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 71, 2, count)


        if i == "2":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 72, 2, count)


        if i == "3":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 73, 2, count)


        if i == "4":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Fail":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 74, 2, count)


        if i == "0":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Fail":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 76, 2, count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 70, 2, count)

        else:
            muc1 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 70, 2, sumarry_case_none)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 76, 2, sumarry_case_none)



    rownum = 9
    #Đặc biệt:
    list_case_nghiemtrong = []
    while (rownum < 1350):
        rownum += 1
        rownum = str(rownum)
        if sheet["L" + rownum].value == "x" and sheet["G" + rownum].value == "Fail":
            print(sheet["A" + rownum].value)
            case_fail = sheet["A" + rownum].value
            list_case_nghiemtrong.append(case_fail)
        rownum = int(rownum)
    print(list_case_nghiemtrong)
    count = len(list_case_nghiemtrong)
    print("Số case fail mức nghiêm trọng: ", count)
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 75, 2, count)


    if var_gpstaxi.modetest == "0":
        case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 70, 2))
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_gpstaxi.modetest == "1":
        case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 71, 2))
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_gpstaxi.modetest == "2":
        case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 72, 2))
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_gpstaxi.modetest == "3":
        case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 73, 2))
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 77, 2, case_fail)
    if var_gpstaxi.modetest == "4":
        case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 74, 2))
        var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 77, 2, case_fail)



def check_casepass():
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 70, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 71, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 72, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 73, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 74, 2, "0")
    var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 75, 2, "0")


    list_casefail = []
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_gpstaxi.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức1: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 81, 2, count)


        if i == "2":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức2: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 82, 2, count)


        if i == "3":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức3: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 83, 2, count)


        if i == "4":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == "Pass":
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức4: ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 84, 2, count)


        if i == "0":
            while (rownum < 1350):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == "Pass":
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            print("Số case trống mức 0 : ", count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 86, 2, count)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 80, 2, count)

        else:
            muc1 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 71, 2))
            muc2 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 72, 2))
            muc3 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 73, 2))
            muc4 = int(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 74, 2))
            sumarry_case_none = muc1 + muc2 + muc3 + muc4
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 80, 2, sumarry_case_none)
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 86, 2, sumarry_case_none)

        if var_gpstaxi.modetest == "0":
            case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 80, 2))
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_gpstaxi.modetest == "1":
            case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 81, 2))
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_gpstaxi.modetest == "2":
            case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 82, 2))
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_gpstaxi.modetest == "3":
            case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 83, 2))
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 87, 2, case_fail)
        if var_gpstaxi.modetest == "4":
            case_fail = str(var_gpstaxi.readData(var_gpstaxi.path_luutamthoi, 'Sheet1', 84, 2))
            var_gpstaxi.writeData(var_gpstaxi.path_luutamthoi, "Sheet1", 87, 2, case_fail)



def change_casenone():
    list_casefail = []
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook.get_sheet_by_name("Checklist")
    rownum = 9
    modetest = ''.join(re.findall(r'\d+', var_gpstaxi.modetest))

    for i in modetest:
        print("i", i)
        if i == "1":
            while (rownum < 500):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G"+rownum].value)
                print(sheet["H"+rownum].value)
                if sheet["H"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var_gpstaxi.checklistpath)
            print("Số case trống mức1: ", count)


        if i == "2":
            while (rownum < 500):
                rownum += 1
                rownum = str(rownum)
                if sheet["I"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var_gpstaxi.checklistpath)
            print("Số case trống mức2: ", count)


        if i == "3":
            while (rownum < 500):
                rownum += 1
                rownum = str(rownum)
                if sheet["J"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var_gpstaxi.checklistpath)
            print("Số case trống mức3: ", count)


        if i == "4":
            while (rownum < 500):
                rownum += 1
                rownum = str(rownum)
                if sheet["K"+rownum].value == "x" and sheet["G"+rownum].value == None:
                    print(sheet["A"+rownum].value)
                    case_fail = sheet["A"+rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var_gpstaxi.checklistpath)
            print("Số case trống mức4: ", count)


        if i == "0":
            while (rownum < 500):
                rownum += 1
                rownum = str(rownum)
                print(sheet["G" + rownum].value)
                print(sheet["H" + rownum].value)
                if (sheet["H" + rownum].value == "x" or sheet["I" + rownum].value == "x" or sheet[
                    "J" + rownum].value == "x" or sheet["K" + rownum].value == "x") and sheet[
                    "G" + rownum].value == None:
                    print(sheet["A" + rownum].value)
                    case_fail = sheet["A" + rownum].value
                    list_casefail.append(case_fail)
                    sheet["G" + rownum] = "Fail"
                rownum = int(rownum)
            print(list_casefail)
            count = len(list_casefail)
            wordbook.save(var_gpstaxi.checklistpath)
            print("Số case trống mức 0 : ", count)




def ModuleTest():
    match = re.search(r'\d+', var_gpstaxi.modetest)
    if not match:
        return None

    module_index = int(match.group())
    return module_index






def Login_Minitor_icon():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(10, 109):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value

        if (mucdo == 0) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level1 != None or level2 != None or level3 != None or level4 != None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level1 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level2 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level3 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level4 == "x"):
            run_cases.append(str(code).strip())

    return run_cases


def Login_Minitor_icon_retest_fail():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(10, 109):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value


        if (mucdo == 0) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] \
                and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level1 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level2 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level3 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level4 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case Fail: {run_cases}")
    return run_cases


def Login_Minitor_icon_retest_none():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(10, 109):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level1 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level2 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level3 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [(code and str(code).startswith("Login")) or (code and str(code).startswith("Minitor"))] and (level4 == "x") and (result == None):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case None: {run_cases}")
    return run_cases





def Minitor_list_vehicle():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(110, 211):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x"):
            run_cases.append(str(code).strip())

    return run_cases


def Minitor_list_vehicle_retest_fail():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(110, 211):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case Fail: {run_cases}")
    return run_cases


def Minitor_list_vehicle_retest_none():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(110, 211):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x") and (result == None):
            run_cases.append(str(code).strip())


    print(f"bắt đầu restest case None: {run_cases}")
    return run_cases







def Minitor_right_Vehicle_Map():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(212, 324):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x"):
            run_cases.append(str(code).strip())

    return run_cases


def Minitor_right_Vehicle_Map_retest_fail():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(212, 324):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case Fail: {run_cases}")
    return run_cases


def Minitor_right_Vehicle_Map_retest_none():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(212, 324):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and (code and str(code).startswith("Minitor")) and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and (code and str(code).startswith("Minitor")) and (level1 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and (code and str(code).startswith("Minitor")) and (level2 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and (code and str(code).startswith("Minitor")) and (level3 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and (code and str(code).startswith("Minitor")) and (level4 == "x") and (result == None):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case None: {run_cases}")
    return run_cases








def Monitor_multiple_Route_Report():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(325, 420):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value

        if (mucdo == 0) and \
                [(code and str(code).startswith("Minitor"))
                 or (code and str(code).startswith("Route"))
                 or (code and str(code).startswith("Report"))] \
                and (level1 != None or level2 != None or level3 != None or level4 != None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level1 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level2 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level3 == "x"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level4 == "x"):
            run_cases.append(str(code).strip())

    return run_cases


def Monitor_multiple_Route_Report_retest_fail():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(325, 420):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and \
                [(code and str(code).startswith("Minitor"))
                 or (code and str(code).startswith("Route"))
                 or (code and str(code).startswith("Report"))] \
                and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level1 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level2 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level3 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level4 == "x") and (result == "Fail"):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case Fail: {run_cases}")
    return run_cases


def Monitor_multiple_Route_Report_retest_none():
    wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
    sheet = wordbook["Checklist"]
    run_cases = []
    mucdo = ModuleTest()

    for rownum in range(325, 420):
        code = sheet[f"A{rownum}"].value
        level1 = sheet[f"H{rownum}"].value
        level2 = sheet[f"I{rownum}"].value
        level3 = sheet[f"J{rownum}"].value
        level4 = sheet[f"K{rownum}"].value
        result = sheet[f"G{rownum}"].value

        if (mucdo == 0) and \
                [(code and str(code).startswith("Minitor"))
                 or (code and str(code).startswith("Route"))
                 or (code and str(code).startswith("Report"))] \
                and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 1) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level1 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 2) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level2 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 3) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level3 == "x") and (result == None):
            run_cases.append(str(code).strip())

        if (mucdo == 4) and [
            (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
            (code and str(code).startswith("Report"))] and (level4 == "x") and (result == None):
            run_cases.append(str(code).strip())

    print(f"bắt đầu restest case None: {run_cases}")
    return run_cases







# level = Retest_case_none()
# print(level)



# def Login_Minitor_icon_Retest_case_fail():
#     wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
#     sheet = wordbook["Checklist"]
#     run_cases = []
#     mucdo = ModuleTest()
#
#     for rownum in range(10, 500):
#         code = sheet[f"A{rownum}"].value
#         level1 = sheet[f"H{rownum}"].value
#         level2 = sheet[f"I{rownum}"].value
#         level3 = sheet[f"J{rownum}"].value
#         level4 = sheet[f"K{rownum}"].value
#         result = sheet[f"G{rownum}"].value
#
#         if (mucdo == 0) and \
#                 [(code and str(code).startswith("Minitor"))
#                  or (code and str(code).startswith("Route"))
#                  or (code and str(code).startswith("Report"))] \
#                 and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == "Fail"):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 1) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level1 == "x") and (result == "Fail"):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 2) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level2 == "x") and (result == "Fail"):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 3) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level3 == "x") and (result == "Fail"):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 4) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level4 == "x") and (result == "Fail"):
#             run_cases.append(str(code).strip())
#
#     return run_cases
#
#
#
# def Login_Minitor_icon_Retest_case_none():
#     wordbook = openpyxl.load_workbook(var_gpstaxi.checklistpath)
#     sheet = wordbook["Checklist"]
#     run_cases = []
#     mucdo = ModuleTest()
#
#     for rownum in range(10, 500):
#         code = sheet[f"A{rownum}"].value
#         level1 = sheet[f"H{rownum}"].value
#         level2 = sheet[f"I{rownum}"].value
#         level3 = sheet[f"J{rownum}"].value
#         level4 = sheet[f"K{rownum}"].value
#         result = sheet[f"G{rownum}"].value
#
#         if (mucdo == 0) and \
#                 [(code and str(code).startswith("Minitor"))
#                  or (code and str(code).startswith("Route"))
#                  or (code and str(code).startswith("Report"))] \
#                 and (level1 != None or level2 != None or level3 != None or level4 != None) and (result == None):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 1) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level1 == "x") and (result == None):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 2) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level2 == "x") and (result == None):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 3) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level3 == "x") and (result == None):
#             run_cases.append(str(code).strip())
#
#         if (mucdo == 4) and [
#             (code and str(code).startswith("Minitor")) or (code and str(code).startswith("Route")) or
#             (code and str(code).startswith("Report"))] and (level4 == "x") and (result == None):
#             run_cases.append(str(code).strip())
#
#     return run_cases





