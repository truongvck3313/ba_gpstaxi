# var_gpstaxi.py
import json
import os
import time
import openpyxl
import logging
from typing import Optional
from get_driver import get_driver
from playwright.async_api import Page, Browser, BrowserContext, Playwright

# ========== Biến toàn cục ==========
p: Optional[Playwright] = None
browser: Optional[Browser] = None
context: Optional[BrowserContext] = None
page: Optional[Page] = None
excelpathdownload: Optional[str] = None
data: dict = {}

# ========== Excel helper ==========
def readData(file: str, sheetName: str, rownum: int, columnno: int):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(row=rownum, column=columnno).value

def writeData(file: str, sheetName: str, rowum: int, columnno: int, data_value):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    sheet.cell(row=rowum, column=columnno).value = data_value
    workbook.save(file)

def writeData_append(file: str, sheetName: str, rowum: int, columnno: int, new_data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    current_value = sheet.cell(row=rowum, column=columnno).value
    if current_value is None:
        current_value = ""
    sheet.cell(row=rowum, column=columnno).value = str(current_value) + str(new_data)
    workbook.save(file)

# ========== Đọc config ==========
def _parse_config_line(line: str):
    """Parse một dòng config dạng '- Key: Value'"""
    try:
        parts = line.split(":", 1)
        if len(parts) < 2:
            return None, None
        key = parts[0].strip()
        value = parts[1].strip().strip('"').strip("'")
        return key, value
    except Exception:
        return None, None

# Khởi tạo biến cấu hình mặc định
modetest = linktest = timerun = logpath = checklistpath = imagepath = datatestpath = ""
uploadpath = path_luutamthoi = excelpath = PATH = moduletest = videopath = ""

# Đọc file_config.txt
config_path = "file_config.txt"
if os.path.exists(config_path):
    with open(config_path, 'r', encoding='utf-8') as f:
        for x in f:
            x_stripped = x.strip()
            if not x_stripped:
                continue
            key, val = _parse_config_line(x_stripped)
            if key is None:
                continue

            if key.startswith("- ModeTest"):
                modetest = val
            elif key.startswith("- LinkTest"):
                linktest = val
            elif key.startswith("- TimeRun"):
                timerun = val
            elif key.startswith("- LogPath"):
                logpath = val
            elif key.startswith("- CheckListPath"):
                checklistpath = val
            elif key.startswith("- ImagePath"):
                imagepath = val
            elif key.startswith("- DataTestPath"):
                datatestpath = val
                if os.path.exists(datatestpath):
                    with open(datatestpath, 'r', encoding='utf-8') as f2:
                        try:
                            data = json.load(f2, strict=False)
                        except Exception:
                            data = {}
                else:
                    data = {}
            elif key.startswith("- UploadPath"):
                uploadpath = val
            elif key.startswith("- ChromeUser"):
                chromeuser = val
            elif key.startswith("- LuuDuLieuTamThoiPath"):
                path_luutamthoi = val
            elif key.startswith("- ExcelPath"):
                excelpath = val
            elif key.startswith("- PathDriver"):
                PATH = val
            elif key.startswith("- ModuleTest"):
                moduletest = val
            elif key.startswith("- ExcelPathDownload"):
                excelpathdownload = val.strip()
            elif key.startswith("- VideoPath"):
                videopath = val

# ========== Khởi tạo Playwright ==========
def init_playwright():
    """
    Khởi tạo Playwright driver.
    Chỉ tạo Chrome test, không dùng Chrome cá nhân.
    """
    global p, browser, context, page

    # try:
    #     # Trước tiên, tắt các Chrome cá nhân đang mở
    #     import subprocess
    #     subprocess.run('taskkill /F /IM chrome.exe /T', shell=True, stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)
    # except Exception:
    #     pass

    try:
        # Khởi tạo driver (get_driver phải trả về Playwright, Browser, Context, Page)
        p, browser, context, page = get_driver(
            headless=False,
            downloads_path=excelpathdownload,
            window_size=(1920, 1480),
            debug_mode=True
        )
        if page:
            try:
                page.set_default_timeout(10000)
            except Exception:
                pass
        print("✅ Playwright driver đã khởi tạo")
    except Exception as e:
        print("❌ Lỗi khi khởi tạo Playwright:", e)

# Khởi tạo nếu chưa có page
# if page is None:
#     init_playwright()





# ========== Restart driver ==========
def restart_driver():
    """
    Restart Playwright driver, giữ chỉ Chrome test.
    """
    global p, browser, context, page
    try:
        if context:
            context.close()
    except Exception:
        pass
    try:
        if browser:
            browser.close()
    except Exception:
        pass
    try:
        if p:
            p.stop()
    except Exception:
        pass

    init_playwright()
    print("🔁 Driver đã restart")











# ========== XPATH ==========
login_user = "//*[@placeholder='Tên đăng nhập']"
login_password = "//*[@placeholder='Mật khẩu']"
login_ghinhodangnhap = "//*[@checked='checked']"
check_danhsachxedangan_xuat_excel = "//*[@id='panelFavourite']//*[@title='Xuất ra Excel']"
icon_bagps = "//*[text()='bagps.vn']"
check_loginsai = "//*[text()='Tên đăng nhập hoặc mật khẩu không hợp lệ.']"
login_iconappstore = "//*[@src='/Images/Login/iconLogin/SVG/photo_logo-IOS.svg']"
check_login_appstore = "//*[text()='BA GPS']"
list_vehicle_vehicle2 = "//*[@id='tblVehicleList']/tbody//*[@vh_online='43C05815_C']/td[2]/div[2]"
dangnhap = "//*[@placeholder='Mật khẩu']"
minitor = "//*[@id='ModuleOnline']/a"
check_goto = "//*[text()='Công ty Bình Anh']"
minitor_vehicle1 = "(//table[contains(@id,'tblVehicleList')])/tbody/tr[2]/td[2]"
minitor_vehicle2 = "(//table[contains(@id,'tblVehicleList')])/tbody/tr[3]/td[2]"
minitor_vehicle3 = "(//table[contains(@id,'tblVehicleList')])/tbody/tr[4]/td[2]"
button_login = "//*[@type='submit']"
RememberMe ="//label[@for='RememberMe']"
login_error = "//*[text()='Tài khoản hoặc mật khẩu không hợp lệ. Vui lòng  đăng nhập lại hoặc liên hệ tới quản trị để được trợ giúp']"
forgot_password = "//*[text()='Quên mật khẩu']"
forgot_user = "//*[@placeholder='Nhập tên tài khoản']"
forgot_phone = "//*[@placeholder='Nhập số điện thoại']"
send_verification_code = "//*[@id='btnGetOTP']"
check_forgot_password1 = "//*[@id='pnlConfirmOTP']/div[1]/div[1]"
check_forgot_password2 = "//*[text()='Mã xác minh được gửi qua zalo của số điện thoại này']"
icon_home = "//*[@src='/images/icons/ic_home.svg']"
icon_sdt = "//*[@class='header']/ul/li[2]"
icon_chplay = "//*[@src='/Content/themes/default/images/photo_logo-android.svg']"
icon_appstore = "//*[@src='/Content/themes/default/images/photo_logo-IOS.svg']"
check_icon_chplay = "//*[text()='BA Taxi']"
check_icon_appstore = "//*[text()='BA TAXI']"
information_and_solutions = "//*[text()='Thông tin giải pháp']"
about_us = "//*[text()='Về chúng tôi']"
network = "//*[text()='Mạng lưới']"
use_help = "//*[text()='Hướng dẫn sử dụng']"
fee_help = "//*[text()='Hướng dẫn đóng phí']"
info_gtvt = "//*[text()='Tin tức ngành GTVT']"
icon_zoom_in = "//*[@title='Zoom in']"
icon_zoom_out = "//*[@title='Zoom out']"
Layers = "//*[@title='Layers']"
name_Layers1 = "(//div[contains(@class,'leaflet-control-layers-base')])/label[1]/div/span"
name_Layers2 = "(//div[contains(@class,'leaflet-control-layers-base')])/label[2]/div/span"
name_Layers3 = "(//div[contains(@class,'leaflet-control-layers-base')])/label[3]/div/span"
close = "//*[text()='Đóng']"
exit = "//*[text()='Thoát']"
close_missing = "//*[@class='btn btn-default btn-close-missing']"
i_close1 = "(//span[contains(@class,'k-icon k-i-close')])[1]"
i_close2 = "(//span[contains(@class,'k-icon k-i-close')])[2]"
i_close3 = "(//span[contains(@class,'k-icon k-i-close')])[3]"
i_close4 = "(//span[contains(@class,'k-icon k-i-close')])[4]"
i_close5 = "(//span[contains(@class,'k-icon k-i-close')])[5]"
i_close6 = "(//span[contains(@class,'k-icon k-i-close')])[6]"
i_close7 = "(//span[contains(@class,'k-icon k-i-close')])[7]"
i_close8 = "(//span[contains(@class,'k-icon k-i-close')])[8]"
i_close9 = "(//span[contains(@class,'k-icon k-i-close')])[9]"
i_close10 = "(//span[contains(@class,'k-icon k-i-close')])[10]"
i_close11 = "(//span[contains(@class,'k-icon k-i-close')])[11]"
radGotoByXNCode = "(//input[contains(@id,'radGotoByXNCode')])"
radGotoByUsername = "(//input[contains(@id,'radGotoByUsername')])"
btnGotoCompany = "(//button[contains(@id,'btnGotoCompany')])"
c08_bca = "(//a[contains(@title,'Danh sách cảnh báo truyền C08 Bộ Công an')])"
panelMissingInfomationBca_wnd_title = "(//span[contains(@id,'panelMissingInfomationBca_wnd_title')])"
txtXNCodeGoto = "//*[@id='txtXNCodeGoto']"

c08_company = "(//ul[contains(@class,'info-missing')])/li[1]/a"
c08_driver = "(//ul[contains(@class,'info-missing')])/li[2]/a"
c08_vehicle = "(//ul[contains(@class,'info-missing')])/li[3]/a"

check_c08_company = "//*[@id='kendo-grid-toolbar-table']//*[text()='Đơn vị kinh doanh vận tải']"
check_c08_vehicle = "//*[@id='kendo-grid-toolbar-table']//*[text()='Loại hình kinh doanh']"
check_c08_driver = "(//div[contains(@id,'DisplayGrid')])//*[text()='Loại bằng']"
icon_refresh = "//*[@title='Làm mới']"
icon_refresh1 = "//*[@id='btn-refresh']"
c08_bca_group = "(//div[contains(@class,'ms-parent')])[1]/button"
c08_bca_group2 = "(//div[contains(@class,'ms-parent')])[1]/div/ul/li[2]"
list_bca_1_1 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[1]"
list_bca_1_2 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[2]"
list_bca_1_3 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[3]"
list_bca_1_4 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[4]"
list_bca_1_5 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[5]"
list_bca_1_6 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[6]"
list_bca_1_7 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[7]"
list_bca_1_7_a = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[7]/a"
list_bca_1_8 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[8]"
list_bca_1_8_input = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[1]/td[8]/input"

list_bca_2_1 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[2]/td[1]"
list_bca_2_2 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[2]/td[2]"
list_bca_2_3 = "(//div[contains(@id,'VehiclesNoTransfer')])//*[@class='k-grid-content']//*[@role='rowgroup']/tr[2]/td[3]"

c08_bca_vehicle = "(//div[contains(@class,'ms-parent')])[2]/button"
c08_bca_vehicle_input = "(//div[contains(@class,'ms-parent')])[2]//*[@type='text']"
c08_bca_vehicle1 = "(//div[contains(@class,'ms-parent')])[2]/div/ul/li[1]/label/input"
cbx_accept = "(//input[contains(@id,'cbx-accept')])"
save_missing = "(//div[contains(@class,'btn btn-primary btn-save-missing')])"
hidden_car_list = "//*[@title='Danh sách xe đã ẩn']"
vehiclesHiddenWindow_wnd_title = "//*[@id='vehiclesHiddenWindow_wnd_title']"
acVehicleSearch = "//*[@id='acVehicleSearch']"
btnAddVehicleHidden = "//*[@id='btnAddVehicleHidden']"
acVehicleHiddenVehicle = "//*[@id='acVehicleHiddenVehicle']"
btnAddVehicleHidden1 = "//*[@id='acVehicleHiddenVehicle_listbox']/li[1]"
radHideTrackingAdd = "//*[@id='radHideTrackingAdd']"
radStopTransferAdd = "//*[@id='radStopTransferAdd']"
btnAddVehicleHidden_resson = "//*[@aria-activedescendant='ddlReasonsAdd_option_selected']/span"
btnAddVehicleHidden_resson1 = "//*[@id='ddlReasonsAdd_listbox']//*[text()='Xe sửa chữa,bảo dưỡng']"
save_value = "//*[@value='Lưu']"
txtNoteAdd = "//*[@id='txtNoteAdd']"
acVehicleSearch_list1 = "//*[@id='acVehicleSearch_listbox']/li[1]"
search_value = "//*[@value='Tìm kiếm']"
k_select4_1_1 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[1]"
k_select4_1_2 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[2]"
k_select4_1_3 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[3]"
k_select4_1_4 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[4]"
k_select4_1_5 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[5]"
k_select4_1_6 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[6]"
k_select4_1_7 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[7]"
k_select4_1_8 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[8]"
k_select4_1_9 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[9]"
k_select4_1_10 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[10]"
k_select4_1_11 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[11]"
# k_select4_1_11_icon = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[1]/td[11]/div/img"
k_select4_1_11_icon = "(//img[contains(@onclick,'HiddenVehicle')])"

k_select4_2_1 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[1]"
k_select4_2_2 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[2]"
k_select4_2_3 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[3]"
k_select4_2_4 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[4]"
k_select4_2_5 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[5]"
k_select4_2_6 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[6]"
k_select4_2_7 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[7]"
k_select4_2_8 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[8]"
k_select4_2_9 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[9]"
k_select4_2_10 = "(//div[contains(@id,'HiddenVehicle')])[1]//*[@role='grid']/tbody/tr[2]/td[10]"
hidden_car_list_resson_icon = "(//span[contains(@aria-owns,'ddlReasons_listbox')])/span//*[@class='k-select']"
hidden_car_list_resson1 = "//*[@id='ddlReasons_listbox']/li[1]"
hidden_car_list_resson2 = "//*[@id='ddlReasons_listbox']//*[text()='Xe sửa chữa,bảo dưỡng']"
HistoryHiddenVehicle1_1 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[1]"
HistoryHiddenVehicle1_2 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[2]"
HistoryHiddenVehicle1_3 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[3]"
HistoryHiddenVehicle1_4 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[4]"
HistoryHiddenVehicle1_5 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[5]"
HistoryHiddenVehicle1_6 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[6]"
HistoryHiddenVehicle1_7 = "(//div[contains(@id,'HistoryHiddenVehicle')])//*[@class='k-grid-content']/table/tbody/tr[1]/td[7]"

hidden_car_list_history_x = "//*[@id='vehiclesHiddenHistoryWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
hidden_car_list_add_new_x = "//*[@id='vehiclesHiddenAddWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
hidden_car_list_status_icon = "(//span[contains(@aria-owns,'ddlSatusHideVehicle_listbox')])/span//*[@class='k-select']"
hidden_car_list_status_minitor = "//*[@id='ddlSatusHideVehicle_listbox']//*[text()='Ẩn trên giám sát']"
hidden_car_list_status_all = "//*[@id='ddlSatusHideVehicle_listbox']/li[1]"
chkStopTransfer = "//*[@id='chkStopTransfer']"
no_permisson = "//*[text()='Không tích tryền']"
search_title = "//*[@title='Search']"
Monitor_multiple_vehicles = "//*[@title='Giám sát nhiều xe']"
Current_Status = "//*[text()='Hiện trạng']"
Measure_distance = "//*[@title='Đo khoảng cách']"
distanceWindow_wnd_title = "//*[@id='distanceWindow_wnd_title']"
radDistanceType = "(//input[contains(@name,'radDistanceType')])[1]"
basic_map = "(//div[contains(@id,'basic-map')])"
marker_icon2 = "(//img[contains(@src,'/images/marker-icon.png')])[2]"
unitsDropdown = "(//a[contains(@id,'unitsDropdown')])"
distanceMajor = "(//span[contains(@id,'distanceMajor')])"
Measure_distance_reset = "(//a[contains(@onclick,'SingletonDistanceManager.getInstance().clear()')])"
distance_km = "(//div[contains(@class,'btn-group btn-group-sm dropdown open')])//*[text()='km']"
distance_m = "(//div[contains(@class,'btn-group btn-group-sm dropdown open')])//*[text()='mét']"
distance_dam = "(//div[contains(@class,'btn-group btn-group-sm dropdown open')])//*[text()='dặm']"
Guide = "//*[@title='Dẫn đường']"
Check_fee = "//*[@title='Tra cước']"
priceCheckWindow_wnd_title = "//*[@id='priceCheckWindow_wnd_title']"
txtEnterKmNumber = "//*[@id='txtEnterKmNumber']"
Check_fee_type_vehicle = "//*[@aria-owns='ddlPriceCheckVehicleType_listbox']//*[@class='k-select']"
ddlPriceCheckVehicleType_listbox_spark_ls = "//*[@id='ddlPriceCheckVehicleType_listbox']//*[text()='Spark LS']"
Check_fee_route = "//*[@aria-owns='ddlPriceCheckRoute_listbox']//*[@class='k-select']"
Check_fee_route_htv_bg = "//*[@id='ddlPriceCheckRoute_listbox']//*[text()='hoàng văn thụ bắc giang']"
Check_fee_payment = "//*[@id='btnPriceCheckSave']"
txtMoneyOfOneWay = "//*[@id='txtMoneyOfOneWay']"
txtMoneyOfTwoWays = "//*[@id='txtMoneyOfTwoWays']"
lblFormula = "//*[@id='lblFormula']"
Make_an_appointment = "//*[@title='Đặt lịch hẹn']"
divEditAppointment_wnd_title = "//*[@id='divEditAppointment_wnd_title']"




CustomerName = "//*[@id='CustomerName']"
PhoneNumber = "//*[@id='PhoneNumber']"
Address = "//*[@id='Address']"
TimeToComeStr = "//*[@id='TimeToComeStr']"
alert_before = "//*[@id='alert_before']/span/span/input[1]"
StartDate = "//*[@id='StartDate']"
IsMultiDays = "//*[@id='IsMultiDays']"
EndDate = "//*[@id='EndDate']"
Mon = "//*[@id='Mon']"
Tue = "//*[@id='Tue']"
Wed = "//*[@id='Wed']"
Fri = "//*[@id='Fri']"
Sat = "//*[@id='Sat']"
Sun = "//*[@id='Sun']"
AllDay = "//*[@id='AllDay']"
Route = "//*[@id='Route']"
RequestVehicles = "//*[@id='RequestVehicles']"
PrivateCodes = "//*[@id='PrivateCodes']"
PaymentTypeId_listbox = "//*[@aria-owns='PaymentTypeId_listbox']//*[@class='k-select']"
PaymentTypeId_listbox1 = "//*[@id='PaymentTypeId_listbox']/li[1]"
CustomerTypeId_listbox = "//*[@aria-owns='CustomerTypeId_listbox']//*[@class='k-select']"
CustomerTypeId_listbox1 = "//*[@id='CustomerTypeId_listbox']/li[1]"

DepositMoney1 = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[2]/input[1]"
TotalMoney1 = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[3]/input[1]"
TotalFee1 = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[4]/input[1]"
DepositMoney = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[2]/input[2]"
TotalMoney = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[3]/input[2]"
TotalFee = "(//span[contains(@class,'k-numeric-wrap k-state-default')])[4]/input[2]"

Note = "//*[@id='Note']"
btnAbort = "//*[@id='btnAbort']"
btnSubmitForm = "//*[@id='btnSubmitForm']"
btnSubmitFormContinue = "//*[@id='btnSubmitFormContinue']"
Appointment_list = "//*[@title='Danh cuộc hẹn']"
panel_title = "//*[@class='panel-title']"
ZoneWarningWindow_wnd_title = "//*[@id='ZoneWarningWindow_wnd_title']"
btnPriceCheckSave = "//*[@id='btnPriceCheckSave']"
Lobby_area_warning_area = "//*[@id='ZoneWarningTabstrip']//*[text()='Vùng']"
Lobby_area_warning_lobby = "//*[@id='ZoneWarningTabstrip']//*[text()='Sảnh']"
Lobby_area_warning_suburb = "//*[@id='ZoneWarningTabstrip']//*[text()='Xe ngoại thành']"
ZoneWarningGrid1_1 = "//*[@id='ZoneWarningGrid']/table/tbody/tr[1]/td[1]"
ZoneWarningGrid1_2 = "//*[@id='ZoneWarningGrid']/table/tbody/tr[1]/td[2]"
ZoneWarningGrid1_3 = "//*[@id='ZoneWarningGrid']/table/tbody/tr[1]/td[3]"
nodata = "//*[text()='Không có dữ liệu']"
Lobby_area_warning_nodata = "//*[@id='HallWarningGrid']//*[text()='Không có dữ liệu']"

ExtramuralWarningGrid1_1 = "//*[@id='ExtramuralWarningGrid']/table/tbody/tr[1]/td[1]"
ExtramuralWarningGrid1_2 = "//*[@id='ExtramuralWarningGrid']/table/tbody/tr[1]/td[2]"
ExtramuralWarningGrid1_3 = "//*[@id='ExtramuralWarningGrid']/table/tbody/tr[1]/td[3]"
btnPriceCheckCancel = "//*[@id='btnPriceCheckCancel']"
Lobby_area_warning = "//*[@title='Cảnh báo vùng sảnh']"
Online_VehicleGroup_listbox = "//*[@aria-owns='Online_VehicleGroup_listbox']//*[@class='k-select']"
Online_VehicleGroup_listbox1 = "//*[@id='Online_VehicleGroup_listbox']/li[1]"
Online_VehicleGroup_listbox2 = "//*[@id='Online_VehicleGroup_listbox']/li[2]"
Online_VehicleStatus_listbox = "//*[@aria-owns='Online_VehicleStatus_listbox']//*[@class='k-select']"
Online_VehicleStatus_listbox1 = "//*[@id='Online_VehicleStatus_listbox']/li[1]"
Online_VehicleStatus_listbox2 = "//*[@id='Online_VehicleStatus_listbox']/li[2]"



ddlSearchProperty_listbox = "//*[@aria-owns='ddlSearchProperty_listbox']//*[@class='k-select']"
ddlSearchProperty_listbox_vehicle = "//*[@id='ddlSearchProperty_listbox']//*[text()='Tìm kiếm xe']"
ddlSearchProperty_listbox_address = "//*[@id='ddlSearchProperty_listbox']//*[text()='Tìm địa chỉ']"
ddlSearchProperty_listbox_landmark = "//*[@id='ddlSearchProperty_listbox']//*[text()='Tìm tên điểm']"
ddlSearchProperty_listbox_coordinates = "//*[@id='ddlSearchProperty_listbox']//*[text()='Tìm tọa độ']"


tblVehicleList1_1 = "//*[@id='tblVehicleList']/tbody/tr[1]/td[1]"
tblVehicleList1_2a = "//*[@id='tblVehicleList']/tbody/tr[1]/td[2]"
tblVehicleList1_2 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[2]"
tblVehicleList1_3 = "//*[@id='tblVehicleList']/tbody/tr[3]/td[3]"
tblVehicleList1_4 = "//*[@id='tblVehicleList']/tbody/tr[4]/td[4]"

tblVehicleList2_1 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[1]"
tblVehicleList2_2 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[2]"
tblVehicleList2_3 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[3]"
tblVehicleList2_4 = "//*[@id='tblVehicleList']/tbody/tr[2]/td[4]"

tblVehicleList3_1 = "//*[@id='tblVehicleList']/tbody/tr[3]/td[1]"
tblVehicleList3_2 = "//*[@id='tblVehicleList']/tbody/tr[3]/td[2]"
tblVehicleList3_3 = "//*[@id='tblVehicleList']/tbody/tr[3]/td[3]"
tblVehicleList3_4 = "//*[@id='tblVehicleList']/tbody/tr[3]/td[4]"

Online_Vehicles_input = "//*[@name='Online_Vehicles_input']"
Online_Vehicles_listbox1 = "//*[@id='Online_Vehicles_listbox']/li[1]"
btnVehicleSearch = "//*[@id='btnVehicleSearch']"
VehicleStatus = "//*[@id='VehicleStatus']"
cbLandMark_input = "//*[@name='cbLandMark_input']"
cbLandMark_listbox1 = "//*[@id='cbLandMark_listbox']/li[1]"
check_Search_landmark = "//*[@class='leaflet-popup-content-wrapper']/div/div/div[1]"
txtPointSearch = "//*[@id='txtPointSearch']"
check_Search_coordinates = "//*[@class='leaflet-popup-content-wrapper']/div/div/div[2]"
hide_detail_vehicle = "//*[@alt='Ẩn chi tiết xe']"
spCurrent = "//*[@id='spCurrent']"
spTotal = "//*[@id='spTotal']"
icon_refresh_new_data = "//*[@title='Cập nhật mới dữ liệu']"
windowCurrentSystem_wnd_title = "//*[@id='windowCurrentSystem_wnd_title']"
currentSystem_listbox = "//*[@aria-owns='currentSystem_listbox']//*[@class='k-select']"
btnSystemStatus = "//*[@id='btnSystemStatus']"
close_button = "//*[@class='leaflet-popup-close-button']"

System_status_have_guests = "//*[@id='currentSystem_listbox']//*[text()='Có khách']"
System_status_no_guests = "//*[@id='currentSystem_listbox']//*[text()='Có khách']"
System_status_lost_singnal = "//*[@id='currentSystem_listbox']//*[text()='Mất tín hiệu']"
System_status_list_vehicle_active = "//*[@id='currentSystem_listbox']//*[text()='Danh sách xe hoạt động']"
System_status_shutdown = "//*[@id='currentSystem_listbox']//*[text()='Tắt máy']"
System_status_too_speed = "//*[@id='currentSystem_listbox']//*[text()='Quá tốc độ']"
System_status_long_stop = "//*[@id='currentSystem_listbox']//*[text()='Dừng đỗ lâu']"
System_status_stopcar_startmachine = "//*[@id='currentSystem_listbox']//*[text()='Dừng xe nổ máy']"
System_status_watch_lock = "//*[@id='currentSystem_listbox']//*[text()='Khóa đồng hồ']"
System_status_lost_watch_connection = "//*[@id='currentSystem_listbox']//*[text()='Mất kết nối đồng hồ']"


ActiveList1_1 = "//*[@id='ActiveList']/table/tbody[1]/tr[1]/td[1]"
ActiveList1_2 = "//*[@id='ActiveList']/table/tbody[1]/tr[1]/td[2]"
ActiveList1_3 = "//*[@id='ActiveList']/table/tbody[1]/tr[1]/td[3]"
ActiveList1_4 = "//*[@id='ActiveList']/table/tbody[1]/tr[1]/td[4]"

distanceWindow_wnd_title_x = "//*[@id='distanceWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
btnExportExcel = "//*[@id='btnExportExcel']"
windowCurrentSystem_wnd_title_x = "//*[@id='windowCurrentSystem_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
btnHelp = "//*[@id='btnHelp']"
helpWindow_wnd_title = "//*[@id='helpWindow_wnd_title']"
helpWindow_wnd_title_x = "//*[@id='helpWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
btnShare = "//*[@id='btnShare']"
shareVehicleWindow_wnd_title = "//*[@id='shareVehicleWindow_wnd_title']"
FromTimeShare = "//*[@id='FromTimeShare']"
ddlShareFunction_listbox = "//*[@aria-owns='ddlShareFunction_listbox']//*[@class='k-select']"
ddlShareFunction_listbox_minitor = "//*[@id='ddlShareFunction_listbox']//*[text()='Giám sát']"
ddlShareFunction_listbox_route = "//*[@id='ddlShareFunction_listbox']//*[text()='Lộ trình']"
ExpiredDateShare = "//*[@id='ExpiredDateShare']"
ExpiredTimeShare = "//*[@id='ExpiredTimeShare']"
ms_parent1 = "(//div[contains(@class,'ms-parent')])[1]/button"
ms_parent2 = "(//div[contains(@class,'ms-parent')])[2]/button"
ms_parent1_li1 = "(//div[contains(@class,'ms-parent')])[1]/div/ul/li[1]/label/input"
ms_parent2_li1 = "(//div[contains(@class,'ms-parent')])[2]/div/ul/li[1]/label/input"
ms_parent2_li2 = "(//div[contains(@class,'ms-parent')])[2]/div/ul/li[2]/label/input"
FromDateShare = "//*[@id='FromDateShare']"
ToTimeShare = "//*[@id='ToTimeShare']"
ToDateShare = "//*[@id='ToDateShare']"
cbxShowLandmark = "//*[@id='cbxShowLandmark']"
cbxAll = "//*[@id='cbxAll']"
create_and_share = "//*[text()='Tạo và chia sẻ']"
coppy = "//*[@title='Sao chép']"
coppy_success = "//*[text()='Sao chép thành công']"
txtLinkShare = "//*[@id='txtLinkShare']"
btnPreview = "//*[@class='btnPreview btn btn-primary']"
remaintime = "//*[@class='left-row remaintime']"
Share_vehicle_minitor_vehicle = "//*[@class='left-row left-row-active']/div[1]"
Share_vehicle_minitor_km = "//*[@class='left-row left-row-active']/div[2]"
Share_vehicle_minitor_time = "//*[@class='left-row left-row-active']/div[3]"
Share_vehicle_minitor_vehicle2 = "//*[@class='panel-info-body']/div[1]/div[2]"
Share_vehicle_minitor_speed = "//*[@class='panel-info-body']/div[2]/div[2]"
Share_vehicle_minitor_address = "//*[@class='panel-info-body']/div[3]/div[2]"

route_remaintime = "//*[@class='left-row-route remaintime']"
Share_vehicle_route_vehicle = "//*[@class='left-panel-router']/div[2]"
Share_vehicle_route_time = "//*[@class='left-panel-router']/div[3]/div[1]/div[2]/div[2]"
Share_vehicle_route_address1 = "//*[@class='left-panel-router']/div[3]/div[2]/div[2]/div[2]"
Share_vehicle_route_address2 = "//*[@class='left-panel-router']/div[3]/div[3]/div[2]/div[2]"
Share_vehicle_minitor_summary = "//*[@class='left-panel-router']/div[3]/div[4]/div[2]/div[2]"

shareVehicleWindow_wnd_title_x = "//*[@id='shareVehicleWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
Online_VehicleStatus_listbox_have_guets = "//*[@id='Online_VehicleStatus_listbox']//*[text()=' -- Có khách']"
tab1_1 = "//*[@id='tab1']/div[1]"
tabs_2 = "//*[@class='tabs']/li[2]"
currentSystem_listbox_have_guests = "//*[@id='currentSystem_listbox']//*[text()='Có khách']"

route_15p_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='15 phút gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_30p_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='30 phút gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_1h_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='1h gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_2h_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='2h gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_4h_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='4h gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_8h_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='8h gần đây']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_inday_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Trong ngày']]  //div[contains(@class,'x-menu-items-event')])[2]"
route_setting_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Tùy chọn']]  //div[contains(@class,'x-menu-items-event')])[2]"


route_15p_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='15 phút gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_30p_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='30 phút gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_1h_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='1h gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_2h_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='2h gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_4h_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='4h gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_8h_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='8h gần đây']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_inday_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Trong ngày']]  //div[contains(@class,'x-menu-items-text')])[2]"
route_setting_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Tùy chọn']]  //div[contains(@class,'x-menu-items-text')])[2]"


icon_route1 = "(//img[contains(@src,'/Images/LoTrinh/LoTrinh')][1])"
divDeleteWrapper_wnd_title = "//*[@id='divDeleteWrapper_wnd_title']"
value_delete = "//*[@value='Xóa']"
divDeleteWrapper_wnd_title_x = "//*[@id='divDeleteWrapper_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"

see_quickly = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Xem nhanh']]  //div[contains(@class,'x-menu-items-event')])[12]"
see_detail_newtab = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Xem chi tiết trên cửa sổ mới']]  //div[contains(@class,'x-menu-items-event')])[12]"
speed = "//*[text()=' Tốc độ : ']"

divDeviceInfo = "//*[@id='divDeviceInfo']/div[1]"
tdDeviceInfo_imei = "//*[@id='tdDeviceInfo']/label[1]"
tdDeviceInfo_number = "//*[@id='tdDeviceInfo']/label[2]"
tdDeviceInfo_liscense_plate = "//*[@id='tdDeviceInfo']/label[3]"
tdDeviceInfo_vin = "//*[@id='tdDeviceInfo']/label[4]"
tdDeviceInfo_time_frist = "//*[@id='tdDeviceInfo']/label[5]"
tdDeviceInfo_time_curent = "//*[@id='tdDeviceInfo']/label[6]"
tdDeviceInfo_gps = "//*[@id='tdSimInfo']/label[1]"
tdDeviceInfo_phone = "//*[@id='tdSim']/label[1]"
tdDeviceInfo_type_phone = "//*[@id='tdSim']/label[2]"
tdDeviceInfo_battery = "//*[@id='tdPower']/label[1]"
tdDeviceInfo_power = "//*[@id='tdPower']/label[2]"
tdDeviceInfo_card = "//*[@id='tdCard']/label[1]"
hide_info_devices = "//*[@alt='Ẩn thông tin thiết bị']"
Minitor_camera = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Giám sát camera']]  //div[contains(@class,'x-menu-items-event')])[1]"
panel_review_title = "//*[@class='panel-reivew-title']"
wrong_date = "//*[@class='wrong-date']"
Subscriber_route = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Nhập thuê bao tuyến']]  //div[contains(@class,'x-menu-items-event')])[2]"



VehicleID_listbox = "//*[@aria-controls='VehicleID_listbox']"
VehicleID_listbox1 = "(//ul[contains(@id,'VehicleID_listbox')])/li[1]"

DriverID_listbox = "//*[@aria-controls='DriverID_listbox']"
DriverID_listbox_1 = "(//ul[contains(@id,'DriverID_listbox')])/li[1]"

StartAddressID_listbox = "//*[@aria-controls='StartAddressID_listbox']"
StartAddressID_listbox2 = "(//ul[contains(@id,'StartAddressID_listbox')])/li[2]"

SubscriberTripID_listbox = "//*[@aria-controls='SubscriberTripID_listbox']"
WayType1 = "(//input[contains(@name,'WayType')])[1]"
WayType2 = "(//input[contains(@name,'WayType')])[2]"
SubscriberTripID_listbox2 = "(//ul[contains(@id,'SubscriberTripID_listbox')])/li[2]"
SubscriberCost = "(//input[contains(@id,'SubscriberCost')])"
FromTime = "(//input[contains(@id,'FromTime')])"
FromDate = "(//input[contains(@id,'FromDate')])[2]"
PickUpKm_click = "(//span[contains(@class,'k-widget k-numerictextbox width100')])[1]"
PickUpKm = "(//input[contains(@id,'PickUpKm')])"
chkHasDropOff = "(//input[contains(@id,'chkHasDropOff')])"
ToDate_dateview = "(//span[contains(@aria-controls,'ToDate_dateview')])/span"
ToTime = "(//input[contains(@id,'ToTime')])"
k_numeric2 = "(//span[contains(@class,'k-widget k-numerictextbox width100')])[2]"
DropOffKm = "(//input[contains(@id,'DropOffKm')])"
k_numeric3 = "(//span[contains(@class,'k-widget k-numerictextbox width100')])[3]"
k_numeric4 = "(//span[contains(@class,'k-widget k-numerictextbox width100')])[4]"
k_numeric5 = "(//span[contains(@class,'k-widget k-numerictextbox width100')])[5]"
RealKm = "(//input[contains(@id,'RealKm')])"
MoneyMeter = "(//input[contains(@id,'MoneyMeter')])"
OverPrice = "(//input[contains(@id,'OverPrice')])"
MoneyReceiver = "(//input[contains(@id,'MoneyReceiver')])"
order_async = "//*[@value='Chọn cuốc đồng bộ']"
btnSubmit = "//*[@id='btnSubmit']"
btnSubmitAndCreateNew = "//*[@id='btnSubmitAndCreateNew']"
btnSubmitAndRedirectTripCost = "//*[@id='btnSubmitAndRedirectTripCost']"
Driver_call1_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Báo nghỉ']]  //div[contains(@class,'x-menu-items-text')])[2]"
Driver_call1_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Báo nghỉ']]  //div[contains(@class,'x-menu-items-event')])[2]"
Driver_call2_name = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Xe hoạt động lại']]  //div[contains(@class,'x-menu-items-text')])[2]"
Driver_call2_button = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Xe hoạt động lại']]  //div[contains(@class,'x-menu-items-event')])[2]"
update_success = "//*[@class='update-success']"
vehiclesHiddenAddWindow_wnd_title = "//*[@id='vehiclesHiddenAddWindow_wnd_title']"
vehiclesHiddenAddWindow_wnd_title_x = "//*[@id='vehiclesHiddenAddWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
tabstrip_2 = "//*[@aria-controls='tabstrip-2']/a"
VehiclePlate_listbox = "//*[@aria-controls='VehiclePlate_listbox']"
VehiclePlate_listbox1 = "//*[@id='VehiclePlate_listbox']/li[1]"
get_data = "//*[@title='Lấy dữ liệu']"
VehiclePlate_input = "//*[@name='VehiclePlate_input']"
divNearestWrapper_wnd_title = "//*[@id='divNearestWrapper_wnd_title']"
ddlSortCarNearest_listbox = "(//span[contains(@aria-owns,'ddlSortCarNearest_listbox')])/span//*[@class='k-select']"
divNearest1_1 = "//*[@id='divNearest']/table/tbody/tr[3]/td[1]"
divNearest1_2 = "//*[@id='divNearest']/table/tbody/tr[3]/td[2]"
divNearest1_3 = "//*[@id='divNearest']/table/tbody/tr[3]/td[3]"
divNearest1_4 = "//*[@id='divNearest']/table/tbody/tr[3]/td[4]"
divNearest1_5 = "//*[@id='divNearest']/table/tbody/tr[3]/td[5]"
divNearest1_6 = "//*[@id='divNearest']/table/tbody/tr[3]/td[6]"
ddlSortCarNearest_listbox1 = "(//ul[contains(@id,'ddlSortCarNearest_listbox')])[1]//*[text()='Khoảng cách  (km)']"
ddlSortCarNearest_listbox2 = "(//ul[contains(@id,'ddlSortCarNearest_listbox')])[1]//*[text()='Doanh thu trong ngày']"
divNearestWrapper_wnd_title_x = "//*[@id='divNearestWrapper_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
Find_forgotten_items = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Tìm đồ khách quên']]  //div[contains(@class,'x-menu-items-event')])[1]"
findCarWindow_wnd_title = "//*[@id='findCarWindow_wnd_title']"
Find_forgotten_items_search = "//*[@onclick='findCarManager.search()']"
rowgroup1_1 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[1]"
rowgroup1_2 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[2]"
rowgroup1_3 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[3]"
rowgroup1_4 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[4]"
rowgroup1_5 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[5]"
rowgroup1_6 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[6]"
rowgroup1_7 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[7]"
rowgroup1_8 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[8]"
rowgroup1_9 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[9]"
rowgroup1_10 = "//*[@id='SearchCarTripGrid']/div[2]/table/tbody/tr[1]/td[10]"
divListDriverNotLoginWrapper_wnd_title = "//*[@id='divListDriverNotLoginWrapper_wnd_title']"
acVehicle = "//*[@id='acVehicle']"
txtTreeView = "//*[@id='txtTreeView']"
List_driver_not_login_group = "(//ul[contains(@class,'k-group')])[2]"
divListDriverNotLogin1_2 = "(//div[contains(@id,'divListDriverNotLogin')])[3]/table[1]/tbody/tr/td[2]"
divListDriverNotLogin2_2 = "(//div[contains(@id,'divListDriverNotLogin')])[3]/table[2]/tbody/tr/td[2]"
acVehicle_listbox1 = "(//ul[contains(@id,'acVehicle_listbox')])/li[1]"
update_data = "//*[@value='Cập nhật dữ liệu']"
fa_fa_times = "//*[@class='fa fa-times']"
ExportExcelDriverNotLogin = "//*[@onclick='FindCar.prototype.ExportExcelDriverNotLogin()']"
Add_a_marker = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Tạo điểm bản đồ']]  //div[contains(@class,'x-menu-items-event')])[1]"
window_wnd_title = "//*[@id='window_wnd_title']"

Name = "//*[@id='Name']"
PrivateName = "//*[@id='PrivateName']"
FK_LandmarkCategoryID_option_selected = "(//span[contains(@aria-activedescendant,'FK_LandmarkCategoryID_option_selected')])//*[@class='k-select']"
park = "(//img[contains(@src,'/Images/Landmark/park.png')])[2]"
IsPolygon1 = "(//input[contains(@id,'IsPolygon')])[1]"
IsPolygon2 = "(//input[contains(@id,'IsPolygon')])[2]"
optionInsert1 = "(//input[contains(@id,'optionInsert')])[1]"
optionInsert2 = "(//input[contains(@id,'optionInsert')])[2]"
txtAddressSuggest = "(//input[contains(@id,'txtAddressSuggest')])"
Longitude = "(//input[contains(@id,'Longitude')])"
Latitude = "(//input[contains(@id,'Latitude')])"
Radius = "(//input[contains(@id,'Radius')])"
IsLandmarkManagement = "(//input[contains(@id,'IsLandmarkManagement')])"
IsVisible = "(//input[contains(@id,'IsVisible')])"
PhoneNumber_input = "(//input[contains(@id,'PhoneNumber')])"
Email = "(//input[contains(@id,'Email')])"
Description = "(//textarea[contains(@id,'Description')])"
HighWayVelocityAllow = "(//input[contains(@id,'HighWayVelocityAllow')])"
LowWayVelocityAllow = "(//input[contains(@id,'LowWayVelocityAllow')])"
IsHallAlert = "(//input[contains(@id,'IsHallAlert')])"
MinVehicle = "(//input[contains(@id,'MinVehicle')])"
MaxVehicle = "(//input[contains(@id,'MaxVehicle')])"
IsLongInAreaAlert = "(//input[contains(@id,'IsLongInAreaAlert')])"
MinuteLongAreaAlert = "(//input[contains(@id,'MinuteLongAreaAlert')])"
addnew_success = "//*[text()='Thêm mới thành công']"
map_icon_park = "//div[contains(@id,'basic-map')]//img[contains(@src,'/Images/Landmark/park.png')]"
Add_a_marker_delete = "//*[@onclick=' return landmarkManager.deleteLandmark(); ']"
Navigation = "(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Dẫn đường']]  //div[contains(@class,'x-menu-items-event')])[1]"
divGaraWrapper_wnd_title = "//*[@id='divGaraWrapper_wnd_title']"

divGara1_1 = "//*[@id='divGara']/table/tbody/tr[3]/td[1]"
divGara1_2 = "//*[@id='divGara']/table/tbody/tr[3]/td[2]"
divGara1_3 = "//*[@id='divGara']/table/tbody/tr[3]/td[3]"
divGaraWrapper_wnd_title_x = "//*[@id='divGaraWrapper_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
divVehiclesInOperateAreaWrapper_wnd_title = "//*[@id='divVehiclesInOperateAreaWrapper_wnd_title']"
divVehiclesInOperateAreaWrapper_wnd_title_x = "//*[@id='divVehiclesInOperateAreaWrapper_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
ddlOperateAreas_listbox = "//*[@id='divInOperateArea']//*[@class='k-select']"
ddlOperateAreas_listbox2 = "(//ul[contains(@id,'ddlOperateAreas_listbox')])[1]/li[2]"
cbInOutOperateAreas = "//*[@id='cbInOutOperateAreas']"

divInOperateArea1_1 = "//*[@id='divInOperateArea']/table/tbody/tr[2]/td[1]"
divInOperateArea1_2 = "//*[@id='divInOperateArea']/table/tbody/tr[2]/td[2]"
divInOperateArea1_3 = "//*[@id='divInOperateArea']/table/tbody/tr[2]/td[3]"
divInOperateArea1_4 = "//*[@id='divInOperateArea']/table/tbody/tr[2]/td[4]"
LandMarkConfig_wnd_title = "//*[@id='LandMarkConfig_wnd_title']"
btnAccept = "//*[@id='btnAccept']"
chkIsShowLabel = "//*[@id='chkIsShowLabel']"
Find_vehicles_in_the_zone = f"(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Tìm xe trong vùng']]  //div[contains(@class,'x-menu-items-event')])[1]"
Find_vehicles_in_the_zone1 = "//*[text()='điểm vùng 1']"
Find_vehicles_in_the_zone2 = "//*[text()='điểm vùng 2']"
Find_vehicles_in_the_zone3 = "//*[text()='điểm vùng 3']"
divAreaWrapper_wnd_title = "//*[@id='divAreaWrapper_wnd_title']"
divArea1_1 = "//*[@id='divArea']/table/tbody/tr[3]/td[1]"
divArea1_2 = "//*[@id='divArea']/table/tbody/tr[3]/td[3]"
divArea1_3 = "//*[@id='divArea']/table/tbody/tr[3]/td[4]"
divArea1_4 = "//*[@id='divArea']/table/tbody/tr[3]/td[5]"
divAreaWrapper_wnd_title_x = "//*[@id='divAreaWrapper_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
Show_hide_boundary = "//*[name()='path' and contains(@class, 'leaflet-interactive')][1]"
Show_hide_boundary_button = f"(//div[contains(@class,'x-menu-item') and .//div[normalize-space()='Ẩn/hiện đường bao']]  //div[contains(@class,'x-menu-items-event')])[1]"
configOnlineDiv_wnd_title = "//*[@id='configOnlineDiv_wnd_title']"


type_map = "//*[@aria-owns='Map_listbox']//*[@class='k-select']"
longitude_input = "//*[@id='txtLongitude1']"
latitude_input = "//*[@id='txtLatitude1']"

zoom_input = "//*[@aria-owns='Zoom_listbox']//*[@class='k-select']"
zoom_input14 = "//*[@id='Zoom_listbox']//*[text()='14']"
btnMapConfigSave = "//*[@id='btnMapConfigSave']"
longitude_click = "(//input[contains(@class,'k-formatted-value k-input')])[1]"
latitude_click = "(//input[contains(@class,'k-formatted-value k-input')])[2]"
Monitor_multiple_vehicles1 = "(//li[contains(@id,'OnlineMultiple')])[1]/a"
check_VehicleStatus = "(//div[contains(@id,'VehicleStatus')])[1]"
remove_circle1 = "(//i[contains(@class,'icon-remove-circle')])[1]"
remove_circle2 = "(//i[contains(@class,'icon-remove-circle')])[2]"
remove_circle3 = "(//i[contains(@class,'icon-remove-circle')])[3]"


Monitor_multiple_x1 = "(//i[contains(@title,'Đóng')])[1]"
Monitor_multiple_x2 = "(//i[contains(@title,'Đóng')])[2]"
Monitor_multiple_x3 = "(//i[contains(@title,'Đóng')])[3]"


Monitor_multiple_zoomin1 = "(//i[contains(@title,'Phóng to')])[1]"
Monitor_multiple_zoomin2 = "(//i[contains(@title,'Phóng to')])[2]"
Monitor_multiple_zoomin3 = "(//i[contains(@title,'Phóng to')])[3]"
CurrentStatus = "(//div[contains(@id,'CurrentStatusZoom')])"
status = "//*[text()='Hiện trạng']"
info_bgt = "//*[text()='Thông tin BGT']"

status_location = "(//article[contains(@id,'tab1')])/div[1]"
status_typevehicle = "(//div[contains(@id,'lstData')])/div[2]/div[2]"
route = "//*[@id='ModuleRoute']/a"
check_route = "//*[text()=' Tốc độ : ']"
route_select = "(//span[contains(@class,'k-widget k-combob')])[2]//*[@class='k-select']"
btnGetData = "//*[@id='btnGetData']"
check_Get_data = "//*[@id='containerRouter']/div[5]/div[2]"
btnSetting = "//*[@id='btnSetting']"
chkDisplayGenieEye_text = "//*[text()='Hiển thị mắt thần']"
chkDisplayGenieEye = "//*[@id='chkDisplayGenieEye']"
scrollFix = "//*[@class='scrollFix']"
chkDisplaySpdMeter = "//*[@id='chkDisplaySpdMeter']"
chkDisplayKm = "//*[@id='chkDisplayKm']"
chkDisplayEngineSts = "//*[@id='chkDisplayEngineSts']"
chkDisplayAcSts = "//*[@id='chkDisplayAcSts']"
chkDisplayLatLong = "//*[@id='chkDisplayLatLong']"
chkDisplayVBGT = "//*[@id='chkDisplayVBGT']"
chkDisplayTemperature = "//*[@id='chkDisplayTemperature']"
chkDisplayCustomer = "//*[@id='chkDisplayCustomer']"
chkDisplayLndMark = "//*[@id='chkDisplayLndMark']"

btnStop = "//*[@id='btnStop']"
btnPlay = "//*[@id='btnPlay']"
btnDecSpeed = "//*[@id='btnDecSpeed']"
btnIncSpeed = "//*[@id='btnIncSpeed']"
btnFastForward = "//*[@id='btnFastForward']"
btnPrint = "//*[@id='btnPrint']"
infoDiv = "//*[@id='infoDiv']/h3"
btnExcel = "//*[@id='btnExcel']"
report = "//*[@id='ModuleReports']/a"
Detailed_trip_report_by_vehicle = "//*[text()='Báo cáo chi tiết cuốc khách theo xe']"
from_money = "//*[text()='Từ tiền']"
btnSearch = "//*[@id='btnSearch']"
DisplayGrid2_1 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[2]/td[1]"
DisplayGrid2_2 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[2]/td[2]"
DisplayGrid2_3 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[2]/td[3]"
DisplayGrid1_1 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[1]/td[1]"
DisplayGrid1_2 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[1]/td[2]"
DisplayGrid1_3 = "(//div[contains(@id,'DisplayGrid')])//*[@role='rowgroup'][2]/tr[1]/td[3]"

Print_report = "//*[@onclick='return report.printReport();']|//*[@onclick='return reportRevenueByVehicle.printReport();']"
check_Print_report = "//*[@id='report_div']/div/table/tbody/tr[2]/td[2]"
export_excel = "//*[text()=' Xuất Excel ']"
reportWindow_wnd_title_x = "//*[@id='reportWindow_wnd_title']/following-sibling::div[contains(@class,'k-window-actions')]//span[contains(@class,'k-i-close')]"
reportWindow_wnd_title = "//*[@id='reportWindow_wnd_title']"
Summary_trip_report_by_vehicle = "//*[text()='Báo cáo tổng hợp cuôc khách theo xe']"
persent_guest = "//*[@id='kendo-grid-toolbar-table']//*[text()='% có khách']"
Speeding_report = "//*[text()='Báo cáo quá tốc độ']"
speeding = "//*[text()='Tốc độ >=']"
btnPause = "//*[@id='btnPause']"












